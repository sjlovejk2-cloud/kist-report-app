"""
AI 보고서 자동 생성기 - Streamlit 웹앱
실행: streamlit run app.py
"""
import streamlit as st
import streamlit.components.v1 as components
import os, re, sys, tempfile, zipfile, hashlib, glob, shutil, subprocess
from io import BytesIO
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from keyword_report import generate_report, build_hwpx, review_design_doc

CONTEXT_FILE = os.path.join(os.path.dirname(__file__), "부서정보.md")
DESIGN_CRITERIA_DIR = os.path.join(os.path.dirname(__file__), "설계검토기준")

st.set_page_config(page_title="AI 보고서 생성기", page_icon="📄", layout="wide")

# Streamlit 우측상단 로딩 픽토그램 숨기기
st.markdown("""
<style>
div[data-testid="stStatusWidget"] { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ── 세션 상태 초기화 ──────────────────────────────────────────
for k, v in [("report_result", None), ("report_error", None),
             ("design_ai_results", None), ("design_ai_error", None), ("design_ai_file_token", None),
             ("design_extracted_fields", {}), ("design_extract_warnings", []),
             ("rag_results", []), ("rag_query", "")]:
    if k not in st.session_state:
        st.session_state[k] = v

# ── 미니게임 HTML ────────────────────────────────────────────
GAME_HTML = """
<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
  *{margin:0;padding:0;box-sizing:border-box;}
  body{background:transparent;display:flex;flex-direction:column;align-items:center;padding:8px;font-family:sans-serif;}
  canvas{border-radius:10px;box-shadow:0 3px 12px rgba(0,0,0,0.18);cursor:pointer;display:block;}
  #info{margin-top:6px;font-size:13px;color:#666;}
</style></head><body>
<canvas id="c" width="680" height="150"></canvas>
<div id="info">SCORE: 0 &nbsp;|&nbsp; BEST: 0</div>
<script>
const canvas=document.getElementById('c'),ctx=canvas.getContext('2d');
const sb=document.getElementById('info');
const W=canvas.width,H=canvas.height;
const GY=112; // ground y
const GRAV=0.72, JV=-14;

let state='idle',score=0,best=0,speed=5,frameN=0,nextObs=90;
let obstacles=[];

const char={x:70,y:GY-44,w:32,h:44,vy:0,
  get bottom(){return this.y+this.h;},
  get grounded(){return this.bottom>=GY;}
};

function jump(){ if(char.grounded) char.vy=JV; }

function drawChar(x,y){
  // 다리
  ctx.fillStyle='#3a6ea8';
  ctx.fillRect(x+6,y+32,8,12); ctx.fillRect(x+18,y+32,8,12);
  // 신발
  ctx.fillStyle='#222';
  ctx.fillRect(x+4,y+42,11,4); ctx.fillRect(x+17,y+42,11,4);
  // 몸통
  ctx.fillStyle='#4a90d9'; ctx.fillRect(x+5,y+18,22,16);
  // 형광조끼
  ctx.fillStyle='#b8e000';
  ctx.fillRect(x+5,y+20,5,12); ctx.fillRect(x+22,y+20,5,12);
  // 팔
  ctx.fillStyle='#4a90d9';
  ctx.fillRect(x-1,y+19,7,5); ctx.fillRect(x+26,y+19,7,5);
  // 손
  ctx.fillStyle='#f5c5a0';
  ctx.fillRect(x-2,y+22,6,4); ctx.fillRect(x+28,y+22,6,4);
  // 목
  ctx.fillStyle='#f5c5a0'; ctx.fillRect(x+12,y+11,8,8);
  // 얼굴
  ctx.fillStyle='#f5c5a0';
  ctx.beginPath(); ctx.arc(x+16,y+9,10,0,Math.PI*2); ctx.fill();
  // 눈
  ctx.fillStyle='#222';
  ctx.beginPath(); ctx.arc(x+12,y+8,2,0,Math.PI*2); ctx.fill();
  ctx.beginPath(); ctx.arc(x+20,y+8,2,0,Math.PI*2); ctx.fill();
  // 입
  ctx.strokeStyle='#a0705a'; ctx.lineWidth=1.5;
  ctx.beginPath(); ctx.arc(x+16,y+11,3.5,0.1,Math.PI-0.1); ctx.stroke();
  // 안전모
  ctx.fillStyle='#FFD700';
  ctx.beginPath(); ctx.ellipse(x+16,y+2,13,7,0,Math.PI,0); ctx.fill();
  ctx.fillRect(x+3,y+1,26,6);
  ctx.fillStyle='#FFC000'; ctx.fillRect(x+1,y+5,30,4);
}

const OBS=['cone','barrier','barrel','sign'];
function makeObs(){
  const t=OBS[Math.floor(Math.random()*OBS.length)];
  const hs={cone:36,barrier:30,barrel:32,sign:40};
  const ws={cone:30,barrier:38,barrel:28,sign:26};
  return{x:W+10,type:t,h:hs[t],w:ws[t]};
}
function drawObs(o){
  const x=o.x,by=GY,h=o.h;
  if(o.type==='cone'){
    ctx.fillStyle='#FF6600';
    ctx.beginPath();ctx.moveTo(x+15,by-h);ctx.lineTo(x,by);ctx.lineTo(x+30,by);ctx.closePath();ctx.fill();
    ctx.fillStyle='white';ctx.fillRect(x+6,by-20,18,4);ctx.fillRect(x+8,by-30,14,4);
  }else if(o.type==='barrier'){
    ctx.fillStyle='#555';ctx.fillRect(x+2,by-h,5,h);ctx.fillRect(x+31,by-h,5,h);
    for(let i=0;i<4;i++){ctx.fillStyle=i%2===0?'#FF2200':'white';ctx.fillRect(x+2+i*9,by-h+3,9,9);}
    ctx.strokeStyle='#333';ctx.lineWidth=1;ctx.strokeRect(x+2,by-h+3,36,9);
  }else if(o.type==='barrel'){
    ctx.fillStyle='#777';ctx.beginPath();ctx.ellipse(x+14,by-h+5,14,4,0,0,Math.PI*2);ctx.fill();
    ctx.fillStyle='#666';ctx.fillRect(x,by-h+5,28,h-5);
    ctx.fillStyle='#777';ctx.beginPath();ctx.ellipse(x+14,by,14,4,0,0,Math.PI*2);ctx.fill();
    ctx.fillStyle='#FFD700';ctx.fillRect(x+2,by-20,24,3);ctx.fillRect(x+2,by-10,24,3);
  }else{
    ctx.fillStyle='#888';ctx.fillRect(x+11,by-h,4,h);
    ctx.fillStyle='#FF8C00';ctx.fillRect(x,by-h,26,20);
    ctx.fillStyle='black';ctx.font='bold 10px sans-serif';ctx.textAlign='center';
    ctx.fillText('공사중',x+13,by-h+8);ctx.fillText('⚠',x+13,by-h+18);ctx.textAlign='left';
  }
}
function hitTest(o){
  const m=6,cx=char.x+m,cy=char.y+m,cw=char.w-m*2,ch=char.h-m*2;
  const ox=o.x+4,oy=GY-o.h+4,ow=o.w-8,oh=o.h-4;
  return cx<ox+ow&&cx+cw>ox&&cy<oy+oh&&cy+ch>oy;
}

let clouds=[{x:180,y:22,s:1},{x:420,y:16,s:0.8},{x:580,y:26,s:1.1}];
function drawCloud(x,y,s){
  ctx.fillStyle='rgba(255,255,255,0.82)';
  ctx.beginPath();ctx.arc(x,y,16*s,0,Math.PI*2);ctx.fill();
  ctx.beginPath();ctx.arc(x+18*s,y-7*s,12*s,0,Math.PI*2);ctx.fill();
  ctx.beginPath();ctx.arc(x+34*s,y,14*s,0,Math.PI*2);ctx.fill();
}
function drawBg(){
  const g=ctx.createLinearGradient(0,0,0,H);
  g.addColorStop(0,'#daeeff');g.addColorStop(1,'#f0f8ff');
  ctx.fillStyle=g;ctx.fillRect(0,0,W,H);
  clouds.forEach(c=>drawCloud(c.x,c.y,c.s));
  ctx.fillStyle='#8BC34A';ctx.fillRect(0,GY+1,W,H-GY-1);
  ctx.fillStyle='#6a9a30';ctx.fillRect(0,GY+1,W,4);
  ctx.strokeStyle='#7aab40';ctx.lineWidth=1;
  ctx.beginPath();ctx.moveTo(0,GY+1);ctx.lineTo(W,GY+1);ctx.stroke();
}

function drawOverlay(l1,l2){
  ctx.fillStyle='rgba(0,0,0,0.35)';ctx.fillRect(0,0,W,H);
  ctx.fillStyle='white';ctx.textAlign='center';
  ctx.font='bold 18px sans-serif';ctx.fillText(l1,W/2,H/2-8);
  ctx.font='13px sans-serif';ctx.fillText(l2,W/2,H/2+14);
  ctx.textAlign='left';
}

function loop(){
  drawBg();
  if(state==='running'){
    clouds.forEach(c=>{c.x-=0.35;if(c.x<-70)c.x=W+40;});
    char.vy+=GRAV; char.y+=char.vy;
    if(char.bottom>=GY){char.y=GY-char.h;char.vy=0;}
    frameN++;
    if(frameN>=nextObs){
      obstacles.push(makeObs());
      nextObs=frameN+80+Math.floor(Math.random()*80);
    }
    obstacles=obstacles.filter(o=>o.x>-60);
    for(const o of obstacles){
      o.x-=speed;
      if(hitTest(o)){state='dead';best=Math.max(best,score);break;}
    }
    score++;
    speed=5+Math.floor(score/500)*0.5;
  }
  obstacles.forEach(drawObs);
  drawChar(char.x,char.y);
  ctx.fillStyle='#333';ctx.font='bold 14px monospace';
  ctx.fillText('SCORE '+String(score).padStart(5,'0'),W-165,20);
  ctx.fillStyle='#888';ctx.font='12px monospace';
  ctx.fillText('BEST  '+String(best).padStart(5,'0'),W-165,36);
  if(state==='idle') drawOverlay('🦺 안전모 점프게임','스페이스바 또는 클릭으로 시작!');
  if(state==='dead') drawOverlay('GAME OVER  💥  '+score+'점','스페이스바 또는 클릭으로 재시작');
  sb.textContent='SCORE: '+score+'  |  BEST: '+best;
  requestAnimationFrame(loop);
}

function handleInput(){
  if(state==='idle'||state==='dead'){
    score=0;speed=5;frameN=0;nextObs=90;obstacles=[];
    char.y=GY-char.h;char.vy=0;state='running';
  }else{ jump(); }
}
document.addEventListener('keydown',e=>{if(e.code==='Space'||e.key===' '){e.preventDefault();handleInput();}});
canvas.addEventListener('click',handleInput);
loop();
</script></body></html>
"""


# ── 설계서 감지/검토 유틸 ───────────────────────────────────────
DESIGN_DOC_KEYWORDS = [
    "설계서", "공사설계서", "내역서", "산출내역", "설계내역", "시방서", "도면",
    "공사명", "공사개요", "공사범위", "공사기간", "공사금액", "예정가격",
    "수량", "단가", "금액", "견적서", "준공", "착공", "안전관리",
    "국가계약법", "지방계약법", "조달청", "원가계산", "간접공사비",
]

DESIGN_CHECKS = [
    {
        "name": "문서 정체성",
        "weight": 15,
        "patterns": ["설계서", "공사설계서", "설계내역", "내역서"],
        "hint": "문서 제목이나 표지에 설계서/내역서 성격이 드러나야 합니다.",
    },
    {
        "name": "공사명 또는 사업명",
        "weight": 12,
        "patterns": ["공사명", "사업명", "용역명", "건명"],
        "hint": "공사명, 사업명, 용역명 또는 건명이 필요합니다.",
    },
    {
        "name": "위치/대상/범위",
        "weight": 12,
        "patterns": ["위치", "장소", "대상", "범위", "공사범위", "공사내용"],
        "hint": "어디에 무엇을 하는 설계인지 범위가 보여야 합니다.",
    },
    {
        "name": "공사기간/일정",
        "weight": 10,
        "patterns": ["공사기간", "기간", "착공", "준공", "납기", "일정"],
        "hint": "착수, 준공, 납기 등 기간 정보가 필요합니다.",
    },
    {
        "name": "금액/예산",
        "weight": 14,
        "patterns": ["공사금액", "예정가격", "합계", "금액", "예산", "부가세"],
        "hint": "총액, 합계, 부가세 포함 여부 등 금액 정보가 필요합니다.",
    },
    {
        "name": "수량/단가/산출 근거",
        "weight": 14,
        "patterns": ["수량", "단가", "산출", "일위대가", "품셈", "규격", "단위"],
        "hint": "수량과 단가, 산출 근거가 있어야 내역 검토가 가능합니다.",
    },
    {
        "name": "시방/품질 기준",
        "weight": 8,
        "patterns": ["시방", "특기시방", "품질", "규격", "기준", "성능"],
        "hint": "자재, 시공, 품질 기준이 빠지면 해석 차이가 생깁니다.",
    },
    {
        "name": "첨부자료",
        "weight": 8,
        "patterns": ["도면", "견적서", "사진", "별첨", "첨부", "도면번호"],
        "hint": "도면, 견적서, 사진, 별첨 목록이 있으면 완성도가 올라갑니다.",
    },
    {
        "name": "안전/법정 검토",
        "weight": 7,
        "patterns": ["안전", "산업안전", "폐기물", "보험료", "환경", "법정"],
        "hint": "안전관리비, 폐기물, 보험료 등 법정성 비용 검토가 필요할 수 있습니다.",
    },
    {
        "name": "계약법 적용 기준",
        "weight": 10,
        "patterns": ["국가계약법", "지방계약법", "국가를 당사자로", "지방자치단체를 당사자로", "계약예규"],
        "hint": "국가계약법/지방계약법 중 적용 기준과 계약예규 적용 여부를 명확히 해야 합니다.",
    },
    {
        "name": "조달청 기준/제비율",
        "weight": 10,
        "patterns": ["조달청", "제비율", "간접공사비", "시설공사 원가계산", "공사원가"],
        "hint": "조달청 시설공사 원가계산 제비율 등 적용 기준과 적용 시점을 확인해야 합니다.",
    },
    {
        "name": "원가계산 구성",
        "weight": 10,
        "patterns": ["재료비", "노무비", "경비", "일반관리비", "이윤", "부가가치세"],
        "hint": "재료비, 노무비, 경비, 일반관리비, 이윤, 부가가치세 구분이 필요합니다.",
    },
    {
        "name": "사후정산/보험료",
        "weight": 8,
        "patterns": ["사후정산", "건강보험료", "연금보험료", "노인장기요양", "산재보험", "고용보험", "퇴직공제"],
        "hint": "보험료와 법정경비의 사후정산 대상 여부를 계약조건과 함께 확인해야 합니다.",
    },
]

RISK_CHECKS = [
    {
        "name": "특정업체 제한 가능성",
        "patterns": ["특정업체", "독점", "단독", "지정제품", "동등 이상 불가", "모델명"],
        "message": "특정 상표·모델·업체로 제한되는 표현이 있으면 경쟁 제한 리스크가 있습니다.",
    },
    {
        "name": "1식 단가 과다 가능성",
        "patterns": ["1식", "일식"],
        "message": "1식 단가는 가능한 세부 수량·단가로 분해해야 안전합니다.",
    },
    {
        "name": "사후정산 누락 가능성",
        "patterns": ["보험료", "안전관리비", "퇴직공제", "사후정산"],
        "message": "보험료·안전관리비 등 정산 조건이 공고/계약조건에 함께 반영되는지 확인해야 합니다.",
        "inverse": True,
    },
]


def _normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


_DOC_TYPE_MAP = {
    "내역서": ["내역서", "내역", "산출내역", "원가계산", "단가표", "물량내역"],
    "시방서": ["시방서", "시방", "특기시방", "표준시방", "공사시방"],
    "도면": ["도면", "평면도", "단면도", "입면도", "상세도", "도면목록", "도면번호", "drawing"],
    "견적서": ["견적서", "견적", "estimate", "quotation", "견적금액"],
}


def detect_doc_type(filename: str, text: str) -> str:
    combined = (filename + " " + text[:400]).lower()
    scores = {
        dtype: sum(1 for kw in kws if kw.lower() in combined)
        for dtype, kws in _DOC_TYPE_MAP.items()
    }
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "기타"


def load_design_criteria_docs() -> list[dict]:
    if not os.path.isdir(DESIGN_CRITERIA_DIR):
        return []

    docs = []
    for name in sorted(os.listdir(DESIGN_CRITERIA_DIR)):
        if not name.lower().endswith(".md"):
            continue
        path = os.path.join(DESIGN_CRITERIA_DIR, name)
        with open(path, "r", encoding="utf-8") as f:
            docs.append({
                "name": name,
                "path": path,
                "content": f.read(),
                "updated_at": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M"),
            })
    return docs


def _extract_uploaded_text(uploaded_file):
    name = uploaded_file.name
    ext = os.path.splitext(name)[1].lower()
    data = uploaded_file.getvalue()
    sheet_names = []

    if ext == ".pdf":
        try:
            import pdfplumber
        except ModuleNotFoundError as exc:
            raise ValueError("PDF 검토에는 pdfplumber가 필요합니다. `pip install pdfplumber` 후 다시 실행해주세요.") from exc
        parts = []
        with pdfplumber.open(BytesIO(data)) as pdf:
            for page in pdf.pages[:20]:
                parts.append(page.extract_text() or "")
        return "\n".join(parts), sheet_names

    if ext in (".xlsx", ".xls"):
        try:
            import pandas as pd
        except ModuleNotFoundError as exc:
            raise ValueError("엑셀 검토에는 pandas/openpyxl이 필요합니다. `pip install pandas openpyxl` 후 다시 실행해주세요.") from exc
        xls = pd.ExcelFile(BytesIO(data))
        sheet_names = xls.sheet_names
        parts = ["시트명: " + ", ".join(sheet_names)]
        for sheet in sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            parts.append(f"\n[{sheet}]")
            parts.append(df.fillna("").astype(str).to_string(index=False, header=False))
        return "\n".join(parts), sheet_names

    if ext in (".txt", ".md", ".csv"):
        return data.decode("utf-8", errors="replace"), sheet_names

    if ext == ".hwpx":
        parts = []
        with zipfile.ZipFile(BytesIO(data), "r") as zf:
            for info in zf.infolist():
                if info.filename.startswith("Contents/") and info.filename.endswith(".xml"):
                    xml = zf.read(info.filename).decode("utf-8", errors="ignore")
                    parts.extend(re.findall(r"<hp:t[^>]*>(.*?)</hp:t>", xml, re.DOTALL))
        text = re.sub(r"<[^>]+>", " ", "\n".join(parts))
        return text, sheet_names

    raise ValueError("지원하지 않는 파일 형식입니다. PDF, XLSX, TXT, MD, CSV, HWPX를 올려주세요.")


def _uploaded_file_token(uploaded_file) -> str:
    data = uploaded_file.getvalue()
    return f"{uploaded_file.name}:{len(data)}:{hashlib.md5(data).hexdigest()}"


def _clean_money_to_int(value: str):
    if not value:
        return None
    s = str(value)
    s = s.replace(",", "").replace("원", "").replace(" ", "")
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else None


def _extract_labeled_text(text: str, labels: list[str], max_len: int = 80):
    for label in labels:
        pattern = rf"{label}\s*[:：]?\s*(.+)"
        m = re.search(pattern, text, re.MULTILINE)
        if m:
            value = m.group(1).strip()
            value = re.split(r"[\r\n]", value)[0].strip()
            return value[:max_len]
    return None


def _extract_labeled_money(text: str, labels: list[str], min_value: int = 1000):
    lines = text.splitlines()
    for label in labels:
        candidates = []
        for line in lines:
            if label not in line:
                continue
            nums = re.findall(r"(?<!\d)(\d{1,3}(?:,\d{3})+|\d+)(?!\d)", line)
            for num in nums:
                value = _clean_money_to_int(num)
                if value is not None and value >= min_value:
                    candidates.append(value)
        if candidates:
            return max(candidates)

        pattern = rf"{label}\s*[:：]?\s*([0-9,]+)\s*원?"
        m = re.search(pattern, text, re.MULTILINE)
        if m:
            value = _clean_money_to_int(m.group(1))
            if value is not None and value >= min_value:
                return value
    return None


def _extract_duration_days(text: str):
    m = re.search(r"공사기간\s*[:：]?\s*(\d+)\s*일", text)
    if m:
        return int(m.group(1))

    m = re.search(r"기간\s*[:：]?\s*(\d+)\s*일", text)
    if m:
        return int(m.group(1))

    m = re.search(r"공사기간\s*[:：]?\s*(\d+)\s*개월", text)
    if m:
        return int(m.group(1)) * 30

    m = re.search(r"기간\s*[:：]?\s*(\d+)\s*개월", text)
    if m:
        return int(m.group(1)) * 30

    return None


def _infer_contract_target(text: str):
    if any(k in text for k in ["용역명", "과업지시서", "제안요청서", "RFP"]):
        return "용역"
    if any(k in text for k in ["물품", "규격서", "사양서", "납품"]):
        return "물품"
    return "공사"


def _infer_construction_type(text: str):
    if any(k in text for k in ["전기", "통신", "소방", "설비", "인테리어", "리모델링"]):
        return "전문공사"
    if any(k in text for k in ["건축", "토목"]):
        return "일반건설공사"
    return "기타"


def _map_to_cost_tool_type(text: str, fields: dict):
    if any(k in text for k in ["전기", "통신", "소방"]):
        return "전기/통신/소방"
    if "설비" in text:
        return "설비(산업설비)"
    if "조경" in text:
        return "조경"
    if "토목" in text:
        return "토목"
    return "건축"


def _map_to_contract_consttype(fields: dict):
    raw = fields.get("construction_type")
    if raw == "일반건설공사":
        return "일반건설공사 (종합공사업 — 건축·토목 등)"
    if raw == "전문공사":
        return "전문공사 (전기·통신·소방·설비 등)"
    return "기타 공사관련 법령"


def extract_design_fields(file_name: str, text: str, sheet_names: list[str], doc_type: str = "기타"):
    raw = text or ""
    one_line = re.sub(r"\s+", " ", raw)

    warnings = []

    project_name = _extract_labeled_text(raw, ["공사명", "사업명", "건명", "용역명"])
    duration_days = _extract_duration_days(raw)

    direct_material_cost = _extract_labeled_money(raw, ["직접재료비", "재료비"], min_value=1000)
    direct_labor_cost = _extract_labeled_money(raw, ["직접노무비", "노무비"], min_value=1000)
    machinery_cost = _extract_labeled_money(raw, ["기계경비"], min_value=1000)
    waste_cost = _extract_labeled_money(raw, ["폐기물처리비"], min_value=1000)
    estimated_price = _extract_labeled_money(raw, ["추정금액", "예정가격", "공사금액", "합계", "견적금액"], min_value=10000)

    contract_target = _infer_contract_target(one_line)
    construction_type = _infer_construction_type(one_line)

    if not project_name:
        project_name = file_name
        warnings.append("공사명을 본문에서 찾지 못해 파일명을 사용했습니다.")

    if estimated_price is None:
        warnings.append("추정금액/합계를 명확히 찾지 못했습니다.")

    if duration_days is None:
        warnings.append("공사기간을 명확히 찾지 못했습니다.")

    fields = {
        "file_name": file_name,
        "doc_type": doc_type,
        "project_name": project_name,
        "contract_target": contract_target,
        "construction_type": construction_type,
        "cost_tool_type": _map_to_cost_tool_type(one_line, {}),
        "contract_consttype": _map_to_contract_consttype({"construction_type": construction_type}),
        "duration_days": duration_days,
        "direct_material_cost": direct_material_cost,
        "direct_labor_cost": direct_labor_cost,
        "machinery_cost": machinery_cost,
        "waste_cost": waste_cost,
        "estimated_price": estimated_price,
        "sheet_names": sheet_names,
        "warnings": warnings,
    }
    return fields


def apply_fields_to_cost_tool(fields: dict):
    if fields.get("project_name"):
        st.session_state["cs_name"] = fields["project_name"]
    if fields.get("duration_days") is not None:
        st.session_state["cs_days_num"] = int(fields["duration_days"])
    if fields.get("cost_tool_type"):
        st.session_state["cs_type"] = fields["cost_tool_type"]
    if fields.get("direct_material_cost") is not None:
        st.session_state["v_mat"] = int(fields["direct_material_cost"])
    if fields.get("direct_labor_cost") is not None:
        st.session_state["v_lab"] = int(fields["direct_labor_cost"])
    if fields.get("machinery_cost") is not None:
        st.session_state["v_mach"] = int(fields["machinery_cost"])
    if fields.get("waste_cost") is not None:
        st.session_state["v_waste"] = int(fields["waste_cost"])

    _d = int(st.session_state.get("cs_days_num", 0) or 0)
    _m = bool(st.session_state.get("cs_ins_manual", False))
    _lock = (0 < _d < 30) and not _m
    st.session_state["r_hi2"] = 0.0 if _lock else 3.595
    st.session_state["r_np2"] = 0.0 if _lock else 4.75
    st.session_state["r_lt2"] = 0.0 if _lock else 13.14


def apply_fields_to_contract_tool(fields: dict):
    if fields.get("contract_target"):
        st.session_state["ct_type"] = fields["contract_target"]

    if fields.get("contract_target") == "공사":
        st.session_state["ct_consttype"] = fields.get("contract_consttype", "기타 공사관련 법령")

    if fields.get("estimated_price") is not None:
        st.session_state["ct_price"] = int(fields["estimated_price"])


def find_small_work_template() -> str | None:
    _candidates = [
        os.path.join(os.path.dirname(__file__), "templates", "승락서,지급각서,하자각서.xlsx"),
        os.path.join(os.path.expanduser("~"), "Desktop", "소액공사 자료", "승락서,지급각서,하자각서.xlsx"),
        "/mnt/c/Users/진광진광/Desktop/소액공사 자료/승락서,지급각서,하자각서.xlsx",
    ]
    _candidates.extend(glob.glob("/mnt/c/Users/*/Desktop/소액공사 자료/승락서,지급각서,하자각서.xlsx"))
    for _path in _candidates:
        if _path and os.path.exists(_path):
            return _path
    return None


def _fmt_date_ymd(_dt) -> str:
    return _dt.strftime("%Y-%m-%d")


def _fmt_period_text(_start, _end) -> str:
    return f"{_fmt_date_ymd(_start)} {_fmt_date_ymd(_end)}"


def _fmt_korean_date_text(_dt) -> str:
    return f"{_dt.year}   년  {_dt.month}월    {_dt.day}일"


def _load_korean_font(size: int, bold: bool = False):
    from PIL import ImageFont

    _candidates = []
    if bold:
        _candidates.extend([
            r"C:\Windows\Fonts\malgunbd.ttf",
            "/mnt/c/Windows/Fonts/malgunbd.ttf",
        ])
    _candidates.extend([
        r"C:\Windows\Fonts\malgun.ttf",
        "/mnt/c/Windows/Fonts/malgun.ttf",
        r"C:\Windows\Fonts\gulim.ttc",
        "/mnt/c/Windows/Fonts/gulim.ttc",
    ])
    for _path in _candidates:
        try:
            if os.path.exists(_path):
                return ImageFont.truetype(_path, size=size)
        except Exception:
            pass
    return ImageFont.load_default()


def _wrap_text_for_draw(_draw, text: str, font, max_width: int) -> list[str]:
    if not text:
        return [""]

    _lines = []
    for _para in str(text).splitlines() or [""]:
        _para = _para.rstrip()
        if not _para:
            _lines.append("")
            continue
        _buf = ""
        for _ch in _para:
            _trial = _buf + _ch
            _bbox = _draw.textbbox((0, 0), _trial, font=font)
            _w = _bbox[2] - _bbox[0]
            if _buf and _w > max_width:
                _lines.append(_buf)
                _buf = _ch
            else:
                _buf = _trial
        if _buf:
            _lines.append(_buf)
    return _lines or [""]


def _draw_wrapped_block(_draw, text: str, font, x: int, y: int, max_width: int, line_gap: int = 10):
    _lines = _wrap_text_for_draw(_draw, text, font, max_width)
    _line_h = (_draw.textbbox((0, 0), "가", font=font)[3] - _draw.textbbox((0, 0), "가", font=font)[1])
    _cursor_y = y
    for _line in _lines:
        _draw.text((x, _cursor_y), _line, font=font, fill="black")
        _cursor_y += _line_h + line_gap
    return _cursor_y


def build_integrity_pledge_pdf(data: dict) -> tuple[bytes, str]:
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from pypdf import PdfReader, PdfWriter
    import pdfplumber

    _template_path = os.path.join(os.path.dirname(__file__), "templates", "청렴계약이행각서_양식.pdf")
    if not os.path.exists(_template_path):
        raise FileNotFoundError("청렴계약이행각서 PDF 양식 파일을 찾지 못했습니다. templates 폴더를 확인하세요.")

    _safe_name = re.sub(r'[\\/:*?"<>|]+', '_', data.get("project_name", "소액공사"))[:80]
    _file_name = f"청렴계약이행각서_{_safe_name}_{data['contract_date'].strftime('%Y%m%d')}.pdf"

    _reader = PdfReader(_template_path)
    _page = _reader.pages[0]
    _page_w = float(_page.mediabox.width)
    _page_h = float(_page.mediabox.height)

    _anchor = None
    with pdfplumber.open(_template_path) as _pdf:
        _words = _pdf.pages[0].extract_words()
        for _w in _words:
            if "업체대표자" in _w.get("text", ""):
                _anchor = _w
                break

    if not _anchor:
        raise RuntimeError("청렴계약이행각서 PDF 양식에서 서약자 위치를 찾지 못했습니다.")

    _font_candidates = [
        ("MalgunGothic", r"C:\Windows\Fonts\malgun.ttf"),
        ("MalgunGothic", "/mnt/c/Windows/Fonts/malgun.ttf"),
        ("Gulim", r"C:\Windows\Fonts\gulim.ttc"),
        ("Gulim", "/mnt/c/Windows/Fonts/gulim.ttc"),
    ]
    _font_name = None
    for _name, _path in _font_candidates:
        try:
            if os.path.exists(_path):
                try:
                    pdfmetrics.getFont(_name)
                except KeyError:
                    pdfmetrics.registerFont(TTFont(_name, _path))
                _font_name = _name
                break
        except Exception:
            pass
    if not _font_name:
        raise RuntimeError("PDF 한글 폰트를 찾지 못했습니다. Windows 한글 폰트를 확인하세요.")

    _overlay_buf = BytesIO()
    _c = canvas.Canvas(_overlay_buf, pagesize=(_page_w, _page_h))

    _line_text = f"서 약 자 : {data.get('ceo_name', '').strip() or data.get('company_name', '').strip() or '서약자'} (인)"
    _font_size = 12

    _x0 = float(_anchor["x0"]) - 2
    _x1 = float(_anchor["x1"]) + 6
    _top = float(_anchor["top"])
    _bottom = float(_anchor["bottom"])
    _y_bottom = _page_h - _bottom
    _box_h = (_bottom - _top) + 8

    _text_w = pdfmetrics.stringWidth(_line_text, _font_name, _font_size)
    _text_x = (_page_w - _text_w) / 2
    _clear_margin = 12
    _clear_x = min(_x0, _text_x - _clear_margin)
    _clear_right = max(_x1, _text_x + _text_w + _clear_margin)

    _c.setFillColorRGB(1, 1, 1)
    _c.rect(_clear_x, _y_bottom - 2, _clear_right - _clear_x, _box_h, fill=1, stroke=0)

    _c.setFillColorRGB(0, 0, 0)
    _c.setFont(_font_name, _font_size)
    _text_y = _page_h - _top - _font_size + 1
    _c.drawString(_text_x, _text_y, _line_text)

    _seal_image_bytes = data.get("seal_image_bytes")
    if _seal_image_bytes:
        try:
            from reportlab.lib.utils import ImageReader
            _seal_reader = ImageReader(BytesIO(_seal_image_bytes))
            _seal_w = 46
            _seal_h = 46
            _seal_x = min(_text_x + _text_w - 18, _page_w - _seal_w - 24)
            _seal_y = _y_bottom - 10
            _c.drawImage(_seal_reader, _seal_x, _seal_y, width=_seal_w, height=_seal_h, mask='auto')
        except Exception:
            pass

    _c.save()
    _overlay_buf.seek(0)

    _overlay_pdf = PdfReader(_overlay_buf)
    _page.merge_page(_overlay_pdf.pages[0])

    _out_buf = BytesIO()
    _writer = PdfWriter()
    _writer.add_page(_page)
    _writer.write(_out_buf)
    _out_buf.seek(0)
    return _out_buf.getvalue(), _file_name


def build_small_work_doc_xlsx(data: dict) -> tuple[bytes, str]:
    _template_path = find_small_work_template()
    if not _template_path:
        raise FileNotFoundError("소액공사 양식 파일을 찾지 못했습니다. Desktop/소액공사 자료 폴더를 확인하세요.")

    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    _wb = openpyxl.load_workbook(_template_path)
    _ws_in = _wb["★자료입력"]
    _ws_accept = _wb["승락서"]
    _ws_pay = _wb["지급각서(7.5%)"]
    _ws_defect = _wb["하자각서(3%)"]

    _contract_date = data["contract_date"]
    _start_date = data["start_date"]
    _end_date = data["end_date"]
    _defect_rate = float(data.get("defect_rate", 0.03) or 0.03)
    _defect_label = data.get("defect_label", "건설업종 3%")
    _defect_period = (data.get("defect_period") or "2년").strip()
    _seal_image_bytes = data.get("seal_image_bytes")

    _ws_in["G2"] = data.get("contract_no", "")
    _ws_in["H2"] = _fmt_date_ymd(_contract_date)
    _ws_in["I2"] = _fmt_period_text(_start_date, _end_date)
    _ws_in["J2"] = data.get("project_name", "")
    _ws_in["K2"] = int(data.get("base_amount", 0) or 0)
    _ws_in["L2"] = int(data.get("contract_amount", 0) or 0)
    _ws_in["M2"] = data.get("contract_method", "수의계약")
    _ws_in["N2"] = data.get("company_name", "")
    _ws_in["O2"] = data.get("biz_no", "")
    _ws_in["P2"] = data.get("ceo_name", "")
    _ws_in["Q2"] = data.get("address", "")

    _ws_pay["C7"] = "=승락서!B7"
    _ws_pay["G7"] = "=승락서!C8"
    _ws_defect["A18"] = _fmt_korean_date_text(_contract_date)
    _ws_defect["B7"] = _defect_period
    _ws_defect["I6"] = f"=I5*{_defect_rate}"
    _ws_defect["B6"] = '=\"일금\"&TEXT(I6,\"[DBNUM4]G/표준\")&\"원정\"'

    for _cell in ["L14", "L15", "L16"]:
        _ws_defect[_cell] = ""
    _mark_map = {
        "건설업종 3%": "L14",
        "조경 5%": "L15",
        "전기, 통신, 소방 등 건설업종 외의 공사 2%": "L16",
    }
    if _defect_label in _mark_map:
        _ws_defect[_mark_map[_defect_label]] = "☑"

    # 원본 양식에 들어있는 기본 직인은 항상 제거합니다.
    # 사용자가 직인 이미지를 업로드한 경우에만 새 직인을 삽입합니다.
    _ws_accept._images = []
    _ws_pay._images = []
    _ws_defect._images = []

    if _seal_image_bytes:
        _png_bytes = BytesIO()
        with PILImage.open(BytesIO(_seal_image_bytes)) as _img:
            _img = _img.convert("RGBA").resize((52, 60))
            _img.save(_png_bytes, format="PNG")
        _png_bytes.seek(0)

        _img1 = XLImage(BytesIO(_png_bytes.getvalue()))
        _ws_accept.add_image(_img1, "D30")

        _img2 = XLImage(BytesIO(_png_bytes.getvalue()))
        _ws_pay.add_image(_img2, "C29")

        _img3 = XLImage(BytesIO(_png_bytes.getvalue()))
        _ws_defect.add_image(_img3, "B25")

    try:
        _wb.calculation.calcMode = "auto"
        _wb.calculation.fullCalcOnLoad = True
        _wb.calculation.forceFullCalc = True
    except Exception:
        pass

    _buf = BytesIO()
    _wb.save(_buf)
    _buf.seek(0)

    _safe_name = re.sub(r'[\\/:*?"<>|]+', '_', data.get("project_name", "소액공사"))[:80]
    _file_name = f"소액공사서류_{_safe_name}_{_contract_date.strftime('%Y%m%d')}.xlsx"
    return _buf.getvalue(), _file_name


def analyze_design_document(file_name: str, text: str, sheet_names: list[str]):
    normalized = _normalize_text(text)
    haystack = f"{file_name} {' '.join(sheet_names)} {normalized}".lower()

    detection_hits = [kw for kw in DESIGN_DOC_KEYWORDS if kw.lower() in haystack]
    detection_score = min(100, int((len(set(detection_hits)) / 8) * 100))
    is_design_doc = detection_score >= 45 or any(word in file_name for word in ("설계서", "내역서", "시방서"))

    checks = []
    total_score = 0
    max_score = sum(item["weight"] for item in DESIGN_CHECKS)
    for item in DESIGN_CHECKS:
        hits = [p for p in item["patterns"] if p.lower() in haystack]
        passed = bool(hits)
        if passed:
            total_score += item["weight"]
        checks.append({
            "항목": item["name"],
            "결과": "충족" if passed else "보완 필요",
            "근거 키워드": ", ".join(hits) if hits else "-",
            "보완 의견": "" if passed else item["hint"],
        })

    risk_flags = []
    for item in RISK_CHECKS:
        hits = [p for p in item["patterns"] if p.lower() in haystack]
        if item.get("inverse"):
            has_fee = any(p.lower() in haystack for p in ["보험료", "안전관리비", "퇴직공제"])
            has_settlement = "사후정산" in haystack or "정산" in haystack
            if has_fee and not has_settlement:
                risk_flags.append({
                    "위험 항목": item["name"],
                    "근거": "보험료/안전관리비 등은 있으나 정산 표현이 부족함",
                    "의견": item["message"],
                })
        elif hits:
            risk_flags.append({
                "위험 항목": item["name"],
                "근거": ", ".join(hits),
                "의견": item["message"],
            })

    quality_score = round(total_score / max_score * 100)
    if quality_score >= 85:
        grade = "양호"
        summary = "필수 구성요소가 대부분 확인됩니다. 세부 수량·단가만 한 번 더 대조하면 좋습니다."
    elif quality_score >= 65:
        grade = "보통"
        summary = "설계서로 볼 수 있으나 일부 핵심 항목 보완이 필요합니다."
    else:
        grade = "미흡"
        summary = "현재 자료만으로는 완성된 설계서로 보기 어렵습니다. 누락 항목 보완이 필요합니다."

    return {
        "is_design_doc": is_design_doc,
        "detection_score": detection_score,
        "detection_hits": detection_hits,
        "quality_score": quality_score,
        "grade": grade,
        "summary": summary,
        "checks": checks,
        "risk_flags": risk_flags,
        "text_length": len(normalized),
        "sheet_count": len(sheet_names),
    }


def make_design_review_text(file_name: str, result: dict) -> str:
    lines = [
        "설계서 검토 결과",
        f"- 파일명: {file_name}",
        f"- 설계서 감지 점수: {result['detection_score']}점",
        f"- 완성도 점수: {result['quality_score']}점",
        f"- 판정: {result['grade']}",
        f"- 종합 의견: {result['summary']}",
        "",
        "세부 점검",
    ]
    for row in result["checks"]:
        lines.append(f"- {row['항목']}: {row['결과']} / 근거: {row['근거 키워드']}")
        if row["보완 의견"]:
            lines.append(f"  보완: {row['보완 의견']}")
    if result.get("risk_flags"):
        lines.extend(["", "위험 신호"])
        for row in result["risk_flags"]:
            lines.append(f"- {row['위험 항목']}: {row['의견']} / 근거: {row['근거']}")
    return "\n".join(lines)


# ── 레이아웃: 좌측 메인(탭) + 우측 원규·지침 검색 패널 ──────────────
_main_col, _right_col = st.columns([2.2, 1], gap="large")

# ══════════════════════════════════════════════════════════════
# 좌측: 기존 탭
# ══════════════════════════════════════════════════════════════
with _main_col:
    tab1, tab2, tab3, tab4 = st.tabs(["📝 보고서 생성", "🧾 설계서 검토", "🧮 업무 도구", "🏢 부서정보 편집"])

    # ── Tab 1: 보고서 생성 ────────────────────────────────────
    with tab1:
        st.subheader("보고서 초안 입력")

        _draft_title = st.text_input(
            "보고서 제목",
            placeholder="예: 2026년 4월 시설공사 추진 현황 보고",
            key="draft_title_input",
        )

        _draft_body = st.text_area(
            "핵심 내용 / 메모",
            placeholder="예: 공사 목적, 진행 현황, 주요 일정, 참석자, 요청사항 등을 자유롭게 적어주세요.",
            height=140,
            key="draft_body_input",
        )

        _gen_col1, _gen_col2 = st.columns([4, 1])
        with _gen_col2:
            generate_btn = st.button("🚀 보고서 초안 생성", type="primary", use_container_width=True)

        st.markdown('입력 예시: "2026년 4월 시설공사 추진 현황 보고"')
        st.markdown("→ 보고 목적, 배경, 현재 상황을 1~3문장 정도로 간단히 적어주세요.")
        st.markdown("→ 날짜, 장소, 참석자, 요청사항, 후속조치가 있으면 함께 적어주면 더 정확한 초안이 생성됩니다.")
        st.markdown("→ 자유 메모 형식으로 작성해도 되며, AI가 보고서 형식으로 정리합니다.")

        if generate_btn and (_draft_title.strip() or _draft_body.strip()):
            _draft_keywords = []
            if _draft_title.strip():
                _draft_keywords.append(_draft_title.strip())
            if _draft_body.strip():
                _draft_keywords.append(_draft_body.strip())

            st.session_state.report_result = None
            st.session_state.report_error = None

            st.info("🤖 AI가 보고서를 작성 중입니다... 아래에서 게임하며 기다리세요! (30초~1분)")
            components.html(GAME_HTML, height=210)

            with st.spinner(""):
                try:
                    st.session_state.report_result = generate_report(_draft_keywords)
                except Exception as e:
                    st.session_state.report_error = str(e)
            st.rerun()

        if st.session_state.report_error:
            st.error(f"보고서 생성 실패: {st.session_state.report_error}")
            st.session_state.report_error = None

        elif st.session_state.report_result is not None:
            report = st.session_state.report_result
            st.success(f"✅ 생성 완료: **{report['title']}**")

            with st.expander("📋 구조 미리보기", expanded=True):
                for sec in report.get("sections", []):
                    st.markdown(f"**□ {sec.get('source','')} — {sec.get('topic','')}**")
                    for pt in sec.get("points", []):
                        st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;○ {pt.get('summary','')}")
                        for d in pt.get("details", []):
                            st.markdown(f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;- {d}")
                    if sec.get("table"):
                        tbl = sec["table"]
                        if tbl.get("headers") and tbl.get("rows"):
                            import pandas as pd
                            df = pd.DataFrame(tbl["rows"], columns=tbl["headers"])
                            st.dataframe(df, hide_index=True, use_container_width=True)

            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"보고서_{ts}.hwpx"
                with tempfile.NamedTemporaryFile(suffix=".hwpx", delete=False) as f:
                    tmp_path = f.name
                build_hwpx(report, tmp_path)
                with open(tmp_path, "rb") as f:
                    hwpx_bytes = f.read()
                os.unlink(tmp_path)
                st.download_button(
                    label="⬇️ HWPX 다운로드",
                    data=hwpx_bytes,
                    file_name=filename,
                    mime="application/octet-stream",
                    type="primary",
                )
            except Exception as e:
                st.error(f"HWPX 변환 실패: {e}")

            st.divider()
            st.caption("🎮 한 판 더? 키워드를 바꿔 다시 생성하면 게임이 시작됩니다.")

    # ── Tab 2: 설계서 검토 ────────────────────────────────────
    with tab2:
        # ── 검토 준비 상태 배지 ─────────────────────────────────
        _criteria_docs = load_design_criteria_docs()
        _db_ok_t2 = False
        _t2_s = {"total_chunks": 0}
        try:
            from kist_rag import db_stats as _t2_db_stats
            _t2_s = _t2_db_stats()
            _db_ok_t2 = _t2_s["ready"]
        except Exception:
            pass

        _bc1, _bc2 = st.columns(2)
        with _bc1:
            if _db_ok_t2:
                st.success(f"🟢 원규·지침 DB 준비 ({_t2_s['total_chunks']:,}개 조항)")
            else:
                st.warning("🔴 원규·지침 DB 미구축 — build_index.py 실행 필요")
        with _bc2:
            if _criteria_docs:
                st.success(f"📋 검토 기준 {len(_criteria_docs)}개 문서 로드됨")
            else:
                st.warning("📋 검토 기준 문서 없음 (설계검토기준/ 폴더 확인)")

        st.divider()

        # ── 파일 업로드 ─────────────────────────────────────────
        _uploaded = st.file_uploader(
            "설계서 파일 업로드 (PDF, XLSX, XLS, TXT, MD, CSV, HWPX)",
            type=["pdf", "xlsx", "xls", "txt", "md", "csv", "hwpx"],
            key="design_uploader",
        )

        if _uploaded:
            try:
                _file_token = _uploaded_file_token(_uploaded)
                if st.session_state.design_ai_file_token != _file_token:
                    st.session_state.design_ai_file_token = _file_token
                    st.session_state.design_ai_results = None
                    st.session_state.design_ai_error = None
                    st.session_state.design_extracted_fields = {}
                    st.session_state.design_extract_warnings = []

                _text, _sheets = _extract_uploaded_text(_uploaded)
                _doc_type = detect_doc_type(_uploaded.name, _text)
                _result = analyze_design_document(_uploaded.name, _text, _sheets)
                _fields = extract_design_fields(_uploaded.name, _text, _sheets, _doc_type)
                st.session_state.design_extracted_fields = _fields
                st.session_state.design_extract_warnings = _fields.get("warnings", [])
                _grade_color = {
                    "양호": "#28a745", "보통": "#fd7e14", "미흡": "#dc3545"
                }.get(_result["grade"], "#6c757d")

                st.markdown(
                    f"<div style='padding:16px;background:#f8f9fa;border-radius:8px;"
                    f"border-left:4px solid {_grade_color};margin-bottom:12px;'>"
                    f"<b>완성도 {_result['quality_score']}점</b> &nbsp; "
                    f"<span style='color:{_grade_color};font-weight:700;'>{_result['grade']}</span>"
                    f"<br><span style='font-size:0.9rem;color:#555;'>{_result['summary']}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

                st.subheader("📋 세부 점검 결과")
                import pandas as pd
                _df_chk = pd.DataFrame([
                    {"항목": c["항목"], "결과": c["결과"],
                     "근거": c["근거 키워드"], "보완의견": c["보완 의견"]}
                    for c in _result["checks"]
                ])
                st.dataframe(_df_chk, hide_index=True, use_container_width=True)

                if _result.get("risk_flags"):
                    st.subheader("⚠️ 위험 신호")
                    st.dataframe(
                        pd.DataFrame(_result["risk_flags"]),
                        hide_index=True, use_container_width=True,
                    )

                st.subheader("🔗 업무 도구로 연결")

                _fields = st.session_state.design_extracted_fields or {}

                if _fields:
                    _summary_rows = [
                        {"항목": "공사명", "값": _fields.get("project_name") or "-"},
                        {"항목": "계약 대상", "값": _fields.get("contract_target") or "-"},
                        {"항목": "공사 구분", "값": _fields.get("construction_type") or "-"},
                        {"항목": "계산기 공사종류", "값": _fields.get("cost_tool_type") or "-"},
                        {"항목": "공사기간(일)", "값": _fields.get("duration_days") if _fields.get("duration_days") is not None else "-"},
                        {"항목": "추정금액", "값": f"{_fields.get('estimated_price', 0):,}원" if _fields.get("estimated_price") is not None else "-"},
                        {"항목": "직접재료비", "값": f"{_fields.get('direct_material_cost', 0):,}원" if _fields.get("direct_material_cost") is not None else "-"},
                        {"항목": "직접노무비", "값": f"{_fields.get('direct_labor_cost', 0):,}원" if _fields.get("direct_labor_cost") is not None else "-"},
                        {"항목": "기계경비", "값": f"{_fields.get('machinery_cost', 0):,}원" if _fields.get("machinery_cost") is not None else "-"},
                        {"항목": "폐기물처리비", "값": f"{_fields.get('waste_cost', 0):,}원" if _fields.get("waste_cost") is not None else "-"},
                    ]

                    st.dataframe(pd.DataFrame(_summary_rows), hide_index=True, use_container_width=True)

                    if st.session_state.design_extract_warnings:
                        for _w in st.session_state.design_extract_warnings:
                            st.caption(f"⚠️ {_w}")

                    _link_col1, _link_col2 = st.columns(2)

                    with _link_col1:
                        if st.button("🧮 계산기로 보내기", use_container_width=True, key="send_to_cost_tool"):
                            apply_fields_to_cost_tool(_fields)
                            st.success("계산기 입력값을 채웠습니다. [업무 도구] 탭에서 확인하세요.")

                    with _link_col2:
                        if st.button("📋 계약판단으로 보내기", use_container_width=True, key="send_to_contract_tool"):
                            apply_fields_to_contract_tool(_fields)
                            st.success("계약판단 입력값을 채웠습니다. [업무 도구] 탭에서 확인하세요.")

                    st.caption("자동 추출값이므로 전송 후 반드시 확인하세요.")
                else:
                    st.info("연결할 추출 정보가 없습니다.")

                st.subheader("🤖 AI 상세 검토")
                if st.button("AI 검토 실행 (30초~1분)", type="primary", key="ai_review_btn"):
                    with st.spinner("AI가 KIST 원규·지침 기반으로 검토 중..."):
                        try:
                            _criteria_texts = [doc.get("content", "") for doc in _criteria_docs]
                            _ai_rev = review_design_doc(_uploaded.name, _doc_type, _text[:6000], _criteria_texts)
                            st.session_state.design_ai_results = _ai_rev
                            st.session_state.design_ai_error = None
                        except Exception as _e:
                            st.session_state.design_ai_error = str(_e)

                if st.session_state.design_ai_results:
                    if isinstance(st.session_state.design_ai_results, dict):
                        _ai = st.session_state.design_ai_results
                        st.markdown(f"**문서 유형 확인:** {_ai.get('doc_type_confirmed', '-')}")
                        st.markdown(f"**완성도 점수:** {_ai.get('completeness_score', '-')}")
                        st.markdown(f"**판정:** {_ai.get('grade', '-')}")
                        st.markdown(f"**종합 의견:** {_ai.get('summary', '-')}")

                        if _ai.get("checks"):
                            st.markdown("**세부 점검**")
                            st.dataframe(pd.DataFrame(_ai["checks"]), hide_index=True, use_container_width=True)

                        if _ai.get("risks"):
                            st.markdown("**위험 요소**")
                            st.dataframe(pd.DataFrame(_ai["risks"]), hide_index=True, use_container_width=True)

                        if _ai.get("missing_companion_docs"):
                            st.markdown("**누락 가능 첨부 문서**")
                            for _doc in _ai["missing_companion_docs"]:
                                st.markdown(f"- {_doc}")
                    else:
                        st.markdown(str(st.session_state.design_ai_results))
                if st.session_state.design_ai_error:
                    st.error(f"AI 검토 오류: {st.session_state.design_ai_error}")

            except ValueError as _e:
                st.error(str(_e))
            except Exception as _e:
                st.error(f"파일 처리 오류: {_e}")
        else:
            st.markdown(
                "<div style='text-align:center;padding:48px 20px;"
                "background:#f8f9fa;border-radius:8px;color:#888;'>"
                "📁 설계서 파일을 업로드하면 자동으로 검토합니다.<br><br>"
                "<small>지원 형식: PDF · XLSX · XLS · TXT · MD · CSV · HWPX</small>"
                "</div>",
                unsafe_allow_html=True,
            )

        # 설계서 검토 탭 사용 가능 상태 — 기존 공사중 오버레이 비활성화
        pass

    # ── Tab 3: 업무 도구 ─────────────────────────────────────
    with tab3:
        _ut1, _ut2, _ut3 = st.tabs(["🧮 조달청 제비율 계산기", "📋 계약 방식 판단", "🗂️ 소액공사 서류 생성기"])

        # ══ 2026.04.13 조달청 건축∙산업환경설비공사 원가계산 간접공사비 적용기준 ══
        # 간접노무비율(직노×%) / 기타경비율((재+노)×%) — (규모키, 기간키): (건축, 산업설비)
        _IL = {  # 간접노무비율
            ('10억미만','6이하'):(17.5,17.5), ('10억미만','7-12'):(17.9,17.9),
            ('10억미만','13-36'):(18.2,18.2), ('10억미만','36초과'):(19.4,19.4),
            ('10-50억','6이하'):(17.4,17.4),  ('10-50억','7-12'):(17.8,17.8),
            ('10-50억','13-36'):(18.1,18.1),  ('10-50억','36초과'):(19.2,19.2),
            ('50-300억','6이하'):(17.5,17.5), ('50-300억','7-12'):(17.9,17.9),
            ('50-300억','13-36'):(18.2,18.2), ('50-300억','36초과'):(19.4,19.4),
            ('300-1000억','6이하'):(17.5,17.5),('300-1000억','7-12'):(17.9,17.9),
            ('300-1000억','13-36'):(18.2,18.2),('300-1000억','36초과'):(19.4,19.4),
            ('1000억이상','6이하'):(17.4,17.4),('1000억이상','7-12'):(17.8,17.8),
            ('1000억이상','13-36'):(18.1,18.1),('1000억이상','36초과'):(19.3,19.3),
        }
        _OE = {  # 기타경비율
            ('10억미만','6이하'):(5.0,5.0),   ('10억미만','7-12'):(5.2,5.2),
            ('10억미만','13-36'):(5.5,5.5),   ('10억미만','36초과'):(6.1,6.1),
            ('10-50억','6이하'):(5.2,5.2),    ('10-50억','7-12'):(5.4,5.4),
            ('10-50억','13-36'):(5.7,5.7),    ('10-50억','36초과'):(6.4,6.4),
            ('50-300억','6이하'):(5.9,5.9),   ('50-300억','7-12'):(6.0,6.0),
            ('50-300억','13-36'):(6.4,6.4),   ('50-300억','36초과'):(7.0,7.0),
            ('300-1000억','6이하'):(6.1,6.1), ('300-1000억','7-12'):(6.3,6.3),
            ('300-1000억','13-36'):(6.6,6.6), ('300-1000억','36초과'):(7.2,7.2),
            ('1000억이상','6이하'):(6.4,6.4), ('1000억이상','7-12'):(6.5,6.5),
            ('1000억이상','13-36'):(6.9,6.9), ('1000억이상','36초과'):(7.5,7.5),
        }
        # 산업안전보건관리비: {분류: [(대상액상한원, 요율%, 기초액천원), ...]}
        _SAFETY = {
            '건축': [(5e8,3.11,0),(50e8,2.28,4325),(800e8,2.37,0),(float('inf'),2.64,0)],
            '토목': [(5e8,3.15,0),(50e8,2.53,3300),(800e8,2.60,0),(float('inf'),2.73,0)],
            '중건설': [(5e8,3.64,0),(50e8,3.05,2975),(800e8,3.11,0),(float('inf'),3.39,0)],
            '특수(조경/전문)': [(5e8,2.07,0),(50e8,1.59,2450),(800e8,1.64,0),(float('inf'),1.78,0)],
        }
        # 환경보전비율 (직접공사비 기준, %)
        _ENV = {
            '건축(주택외/연구시설)': 0.5, '건축(주택신축)': 0.3,
            '토목(도로/교량/터널)': 0.9, '토목(하천/기타)': 0.8,
            '토목(택지개발)': 0.6, '토목(상하수도)': 0.5,
            '조경/전문/기타': 0.3,
        }

        def _get_safety_fee(대상액, 분류='건축'):
            for 상한, 율, 기초 in _SAFETY.get(분류, _SAFETY['건축']):
                if 대상액 < 상한:
                    return 대상액 * (율 / 100) + 기초 * 1000
            return 대상액 * (_SAFETY.get(분류, _SAFETY['건축'])[-1][1] / 100)

        def _get_size_key(직접공사비):
            if 직접공사비 < 10e8: return '10억미만'
            elif 직접공사비 < 50e8: return '10-50억'
            elif 직접공사비 < 300e8: return '50-300억'
            elif 직접공사비 < 1000e8: return '300-1000억'
            else: return '1000억이상'

        def _num_to_kor(n):
            """숫자를 한글 금액으로 변환"""
            n = int(round(n))
            if n == 0: return '영원정'
            units = ['', '만', '억', '조']
            parts = []
            temp = n
            for u in units:
                parts.append((temp % 10000, u))
                temp //= 10000
                if temp == 0: break
            result = ''
            for v, u in reversed(parts):
                if v > 0:
                    result += f'{v}{u}'
            return f'일금{result}원정'

        # ── 제비율 계산기 ──────────────────────────────────────
        with _ut1:
            st.caption("📅 적용기준: 2026.04.13 입찰공고분부터 (조달청 건축∙산업환경설비공사 원가계산 간접공사비 적용기준)")
            if st.session_state.get("cs_name"):
                st.caption("📥 설계서 검토 탭에서 전달된 값이 자동 반영될 수 있습니다.")

            # ① 공사 정보
            st.subheader("① 공사 정보")

            # ─ 사회보험료 자동화 초기화 (session_state 기본값) ─
            for _k, _dv in [("r_hi2", 3.595), ("r_np2", 4.75), ("r_lt2", 13.14)]:
                if _k not in st.session_state:
                    st.session_state[_k] = _dv

            def _sync_insurance():
                """공사기간(일) or 수동입력 변경 시 사후정산 보험료 자동 처리"""
                _d = int(st.session_state.get("cs_days_num", 0) or 0)
                _m = bool(st.session_state.get("cs_ins_manual", False))
                _lock = (0 < _d < 30) and not _m
                st.session_state["r_hi2"] = 0.0 if _lock else 3.595
                st.session_state["r_np2"] = 0.0 if _lock else 4.75
                st.session_state["r_lt2"] = 0.0 if _lock else 13.14

            # 보험료 잠금 여부 (30일 미만 & 수동입력 미체크)
            _ins_days = int(st.session_state.get("cs_days_num", 0) or 0)
            _ins_lock = (0 < _ins_days < 30) and not st.session_state.get("cs_ins_manual", False)

            _ni1, _ni2, _ni3 = st.columns([3, 1, 1])
            with _ni1:
                _공사명 = st.text_input("공사명", placeholder="예: 0000연구단(0000호) 실험실공사(설비)", key="cs_name")
            with _ni2:
                _공사일자 = st.date_input("작성일자", value=datetime.today(), key="cs_date")
            with _ni3:
                st.number_input(
                    "공사기간(일)", min_value=0, value=0, step=1, format="%d",
                    key="cs_days_num", on_change=_sync_insurance,
                    help="30일 미만: 건강보험료·국민연금·장기요양 자동 미계상(사후정산 대상)")

            _nt1, _nt2, _nt3 = st.columns(3)
            with _nt1:
                _공사종류 = st.selectbox("공사종류",
                    ["건축", "설비(산업설비)", "전기/통신/소방", "조경", "토목"], key="cs_type")
            with _nt2:
                _규모선택 = st.selectbox("직접공사비 규모 (요율 자동 적용)",
                    ["자동", "10억 미만", "10-50억", "50-300억", "300-1000억", "1000억 이상"], key="cs_size")
            with _nt3:
                _기간선택 = st.selectbox("공사기간",
                    ["6개월 이하", "7~12개월", "13~36개월", "36개월 초과"], key="cs_dur")

            _소액여부 = st.checkbox(
                "소액공사 적용 (간접노무비·기타경비 제외)", value=False, key="cs_small",
                help="소액공사는 간접노무비와 기타경비를 계상하지 않습니다")

            # ② 공사비 입력
            st.subheader("② 공사비 입력")
            _n1, _n2, _n3 = st.columns(3)
            with _n1:
                _직접재료비 = st.number_input("가. 직접재료비 (원)", min_value=0, value=0, step=100_000, format="%d", key="v_mat")
            with _n2:
                _직접노무비 = st.number_input("나. 직접노무비 (원)", min_value=0, value=0, step=100_000, format="%d", key="v_lab")
            with _n3:
                _기계경비   = st.number_input("다. 기계경비 (원)",   min_value=0, value=0, step=100_000, format="%d", key="v_mach")

            _폐기물처리비 = st.number_input("폐기물처리비 (원, 별도 계상 시)", min_value=0, value=0, step=10_000, format="%d", key="v_waste")

            # ③ 요율 확인/수정
            st.subheader("③ 적용 요율 확인 / 수정")

            # 공사종류·규모·기간으로 기본 요율 결정
            _기간키 = {'6개월 이하':'6이하','7~12개월':'7-12','13~36개월':'13-36','36개월 초과':'36초과'}[_기간선택]
            _안전분류 = {
                '건축':'건축', '설비(산업설비)':'건축',
                '전기/통신/소방':'건축', '조경':'특수(조경/전문)', '토목':'토목'
            }[_공사종류]
            _환경분류 = {
                '건축':'건축(주택외/연구시설)', '설비(산업설비)':'건축(주택외/연구시설)',
                '전기/통신/소방':'건축(주택외/연구시설)',
                '조경':'조경/전문/기타', '토목':'토목(하천/기타)'
            }[_공사종류]
            _is산업설비 = '설비' in _공사종류  # 간접노무비/기타경비율 인덱스 (0=건축, 1=산업설비)

            # 규모키 (직접공사비 입력값 기반 or 수동 선택)
            _규모맵 = {'10억 미만':'10억미만','10-50억':'10-50억','50-300억':'50-300억',
                       '300-1000억':'300-1000억','1000억 이상':'1000억이상'}
            _규모키_기본 = _get_size_key(_직접재료비 + _직접노무비 + _기계경비)
            _규모키 = _규모맵.get(_규모선택, _규모키_기본) if _규모선택 != '자동' else _규모키_기본

            _il_default = _IL.get((_규모키, _기간키), (17.5, 17.5))[1 if _is산업설비 else 0]
            _oe_default = _OE.get((_규모키, _기간키), (5.0, 5.0))[1 if _is산업설비 else 0]

            with st.expander(f"📊 현재 적용 요율 ({_공사종류} / {_규모키} / {_기간선택})", expanded=True):
                _rv1, _rv2, _rv3 = st.columns(3)
                with _rv1:
                    st.markdown("**간접공사비**")
                    _r_간접노무 = st.number_input(f"간접노무비율 (직노×%)",
                        value=float(_il_default) if not _소액여부 else 0.0,
                        step=0.1, format="%.1f", key="r_il",
                        disabled=_소액여부)
                    _r_기타경비 = st.number_input(f"기타경비율 ((재+노)×%)",
                        value=float(_oe_default) if not _소액여부 else 0.0,
                        step=0.1, format="%.1f", key="r_oe",
                        disabled=_소액여부)
                with _rv2:
                    # 헤더 + 수동 입력 체크박스
                    _rv2h, _rv2c = st.columns([2, 1])
                    with _rv2h:
                        st.markdown("**사회보험료** (2026 기준)")
                    with _rv2c:
                        st.checkbox("수동 입력", value=False, key="cs_ins_manual",
                            on_change=_sync_insurance,
                            help="체크 시 공사기간 자동화 무시 — 요율 직접 입력 가능")

                    # 사후정산 대상 3개 (공사기간 30일 미만 시 자동 0 / 비활성화)
                    _r_건강보험 = st.number_input("건강보험료 (직노×%)",
                        min_value=0.0, step=0.001, format="%.3f",
                        key="r_hi2", disabled=_ins_lock)
                    _r_연금     = st.number_input("국민연금 (직노×%)",
                        min_value=0.0, step=0.01,  format="%.2f",
                        key="r_np2", disabled=_ins_lock)
                    _r_고용     = st.number_input("고용보험료 (노×%)",
                        value=1.05,   step=0.01,   format="%.2f", key="r_ei2")
                    _r_산재     = st.number_input("산재보험료 (노×%)",
                        value=3.56,   step=0.01,   format="%.2f", key="r_wi2")
                    _r_장기요양 = st.number_input("장기요양 (건강보험료×%)",
                        min_value=0.0, step=0.01,  format="%.2f",
                        key="r_lt2",  disabled=_ins_lock)

                    # 30일 미만 안내 메시지
                    if _ins_lock:
                        st.markdown(
                            "<div style='margin-top:6px;padding:7px 10px;"
                            "background:#fff5f5;border-left:3px solid #e03131;"
                            "border-radius:4px;font-size:12px;color:#c92a2a;line-height:1.5;'>"
                            "안내: 공사기간 30일 미만으로<br>"
                            "사후정산 보험료(건강·연금·장기요양)가<br>"
                            "미계상(0%) 처리되었습니다.</div>",
                            unsafe_allow_html=True)
                with _rv3:
                    st.markdown("**일반관리비·이윤**")
                    _r_일반관리 = st.number_input("일반관리비율 (순공사비×%, 상한6%)",
                        value=6.0, step=0.1, format="%.1f", key="r_adm2", max_value=6.0)
                    _r_이윤     = st.number_input("이윤율 ((노+경+일)×%, 상한15%)",
                        value=15.0, step=0.1, format="%.1f", key="r_pft2", max_value=15.0)
                    st.markdown("**산업안전보건관리비**")
                    _안전입력방식 = st.radio("입력 방식",
                        ["자동 (대상액×요율+기초액)", "금액 직접 입력"],
                        horizontal=False, key="r_safe_mode")
                    _안전직접 = 0
                    if _안전입력방식 == "금액 직접 입력":
                        _안전직접 = st.number_input("산업안전보건관리비 (원)",
                            min_value=0, value=0, step=10_000, key="r_safe_amt")
                    st.markdown("**환경보전비**")
                    _환경입력방식 = st.radio("환경보전비",
                        [f"자동 ({_ENV[_환경분류]}% · {_환경분류})", "직접 입력", "미계상"],
                        horizontal=False, key="r_env_mode")
                    _환경직접 = 0.0
                    if _환경입력방식 == "직접 입력":
                        _환경직접 = st.number_input("환경보전비율 (%)", value=0.5, step=0.1, format="%.1f", key="r_env_val")

            # ④ 계산
            if st.button("💰 계산하기", type="primary", use_container_width=True, key="calc_btn2"):
                import pandas as pd

                _재료비      = float(_직접재료비)
                _간접노무비  = _재료비 * 0 + _직접노무비 * (_r_간접노무 / 100) if not _소액여부 else 0.0
                _노무비      = _직접노무비 + _간접노무비
                _기타경비_v  = (_재료비 + _노무비) * (_r_기타경비 / 100) if not _소액여부 else 0.0
                _대상액      = _재료비 + _직접노무비      # 안전관리비 기준

                # 산업안전보건관리비
                if _안전입력방식 == "금액 직접 입력":
                    _안전관리비 = float(_안전직접)
                else:
                    _안전관리비 = _get_safety_fee(_대상액, _안전분류)

                # 환경보전비
                _직접공사비 = _재료비 + _노무비 + float(_기계경비)
                if _환경입력방식 == "미계상":
                    _환경보전비 = 0.0
                elif _환경입력방식 == "직접 입력":
                    _환경보전비 = _직접공사비 * (_환경직접 / 100)
                else:
                    _환경보전비 = _직접공사비 * (_ENV[_환경분류] / 100)

                # 사회보험료 (2026기준: 건강·연금은 직노 기준, 고용·산재는 노무비 기준)
                _건강보험료  = _직접노무비 * (_r_건강보험 / 100)
                _연금보험료  = _직접노무비 * (_r_연금     / 100)
                _고용보험료  = _노무비     * (_r_고용     / 100)
                _산재보험료  = _노무비     * (_r_산재     / 100)
                _장기요양    = _건강보험료 * (_r_장기요양 / 100)

                _경비 = (float(_기계경비) + _기타경비_v + _안전관리비 + _환경보전비
                        + _건강보험료 + _연금보험료 + _고용보험료 + _산재보험료
                        + _장기요양)

                _순공사비    = _재료비 + _노무비 + _경비
                _일반관리비  = _순공사비 * (_r_일반관리 / 100)
                _소계        = _순공사비 + _일반관리비
                _이윤        = (_노무비 + _경비 + _일반관리비) * (_r_이윤 / 100)
                _폐기물      = float(_폐기물처리비)
                _계          = _소계 + _이윤 + _폐기물
                _부가세      = _계 * 0.1
                _합계        = _계 + _부가세

                def _fmt(n): return f"{int(round(n)):,}"

                # 결과 표시 (설계서용지 구조와 동일)
                st.subheader("④ 계산 결과")
                _result_rows = [
                    (True,  False, "1",  "순공사비",          "",                                         _순공사비),
                    (False, False, "",   "  가. 직접재료비",   "",                                         _재료비),
                    (False, False, "",   "  나. 노무비",       "",                                         _노무비),
                    (False, False, "",   "    직접노무비",     "",                                         _직접노무비),
                    (False, _소액여부, "", f"    간접노무비",  f"직노의 {_r_간접노무:.1f}%",               _간접노무비),
                    (False, False, "",   "  다. 경비",         "",                                         _경비),
                    (False, False, "",   "    기계경비",       "",                                         float(_기계경비)),
                    (False, _소액여부, "", "    기타경비",     f"(재+노)의 {_r_기타경비:.1f}%",            _기타경비_v),
                    (False, False, "",   "    산재보험료",     f"노의 {_r_산재:.2f}%",                     _산재보험료),
                    (False, False, "",   "    고용보험료",     f"노의 {_r_고용:.2f}%",                     _고용보험료),
                    (False, False, "",   "    건강보험료",     f"직노의 {_r_건강보험:.3f}%",               _건강보험료),
                    (False, False, "",   "    국민연금",       f"직노의 {_r_연금:.2f}%",                   _연금보험료),
                    (False, False, "",   "    장기요양보험료", f"건강보험료의 {_r_장기요양:.2f}%",         _장기요양),
                    (False, False, "",   "    산업안전보건관리비", f"{_안전분류} / 대상액 기준",           _안전관리비),
                    (False, False, "",   "    환경보전비",     f"{_환경분류}",                             _환경보전비),
                    (True,  False, "2",  "일반관리비",         f"순공사비의 {_r_일반관리:.0f}%",           _일반관리비),
                    (True,  False, "",   "    소   계",        "",                                         _소계),
                    (True,  False, "3",  "이   윤",            f"(노+경+일관)의 {_r_이윤:.0f}%",          _이윤),
                    (False, False, "4",  "폐기물처리비",       "",                                         _폐기물),
                    (True,  False, "",   "    계",             "",                                         _계),
                    (False, False, "5",  "부가가치세",         "계의 10%",                                 _부가세),
                    (True,  True,  "",   "    합   계",        "",                                         _합계),
                ]

                for _bold, _total, _no, _name, _spec, _val in _result_rows:
                    _bg = "#1a73e8" if _total else ("#f0f4ff" if _bold else "white")
                    _fc = "white" if _total else ("#1a73e8" if _bold else "#333")
                    _fw = "700" if (_bold or _total) else "400"
                    _val_display = _fmt(_val) if _val > 0 else ("0" if _val == 0 and (_bold or _total or _no != "") else "")
                    st.markdown(
                        f"<div style='display:flex;padding:5px 12px;background:{_bg};"
                        f"border-radius:5px;margin-bottom:2px;gap:8px;'>"
                        f"<span style='width:20px;color:{_fc};font-size:13px;'>{_no}</span>"
                        f"<span style='flex:2;color:{_fc};font-weight:{_fw};font-size:13px;'>{_name}</span>"
                        f"<span style='flex:2;color:#888;font-size:12px;'>{_spec}</span>"
                        f"<span style='flex:1;text-align:right;color:{_fc};font-weight:{_fw};font-size:13px;'>"
                        f"{_val_display}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )

                st.caption(f"합계: {_num_to_kor(_합계)} ({_fmt(_합계)}원)")
                st.caption("⚠️ 참고용 계산기입니다. 실제 적용 요율은 조달청 고시 및 KIST 계약팀에 확인하세요.")

                # ⑤ 설계서용지 Excel 출력
                st.subheader("⑤ 설계서용지 Excel 다운로드")
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                    from openpyxl.utils import get_column_letter
                    from io import BytesIO

                    _wb = openpyxl.Workbook()
                    _ws = _wb.active
                    _ws.title = "설계서용지"

                    # 열 너비 설정 (번호, 명칭, 규격, 단위, 수량, 단가, 금액, 비고)
                    for _ci, _cw in enumerate([6, 28, 22, 6, 7, 14, 16, 12], 1):
                        _ws.column_dimensions[get_column_letter(_ci)].width = _cw

                    _thin = Side(style='thin')
                    _bdr  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                    _bdr_top = Border(top=_thin)

                    def _ws_set(r, c, val, bold=False, align='left', size=10, fill=None, border=None, merge_to=None):
                        cell = _ws.cell(row=r, column=c, value=val)
                        cell.font = Font(name='맑은 고딕', bold=bold, size=size)
                        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                        if fill:
                            cell.fill = PatternFill('solid', fgColor=fill)
                        if border:
                            cell.border = border
                        if merge_to:
                            _ws.merge_cells(start_row=r, start_column=c, end_row=merge_to[0], end_column=merge_to[1])
                        return cell

                    # 제목
                    _ws.row_dimensions[1].height = 24
                    _ws_set(1, 1, "설 계 서 용 지", bold=True, size=14, align='center', merge_to=(1,8))
                    # 일자
                    _dt_str = _공사일자.strftime("%Y. %m. %d")
                    _ws_set(2, 1, f"일자 : {_dt_str}", size=10, merge_to=(2,8))
                    # 공사명
                    _ws_set(3, 1, f"공 사 명    :   {_공사명 or '(공사명 미입력)'}", size=10, merge_to=(3,8))
                    # 금액
                    _ws.row_dimensions[4].height = 6
                    _ws_set(5, 1, "금     액    :", size=10)
                    _ws_set(5, 2, _num_to_kor(_합계), bold=True, size=10, merge_to=(5,7))
                    _c5g = _ws_set(5, 8, "=G29", bold=True, size=10, align='right')
                    _c5g.number_format = '#,##0'
                    _ws.row_dimensions[6].height = 6

                    # 헤더
                    _ws.row_dimensions[7].height = 22
                    _hdr = ["번호", "명    칭", "규    격", "단위", "수    량", "단    가", "금    액", "비   고"]
                    for _ci, _h in enumerate(_hdr, 1):
                        _ws_set(7, _ci, _h, bold=True, align='center', size=10, fill='D9E1F2', border=_bdr)

                    # 데이터 행 정의 (G열 수식 연동 / 직접 입력값은 정수 고정)
                    # 행 번호: 순공사비=8, 재료비=9, 노무비=10, 직노=11, 간노=12,
                    #          경비=13, 기계=14, 기타=15, 산재=16, 고용=17,
                    #          건강=18, 연금=19, 장기요양=20, 안전=21, 환경=22,
                    #          일관=23, 소계=24, 이윤=25, 폐기물=26, 계=27, 부가세=28, 합계=29
                    _f_간접 = f"=G11*{_r_간접노무}/100" if not _소액여부 else 0
                    _f_기타 = f"=(G9+G10)*{_r_기타경비}/100" if not _소액여부 else 0
                    _sheet_rows = [
                        ("1",  "순공사비",              "",                                                   "=G9+G10+G13",                     True),
                        ("",   "가. 직접 재료비",        "",                                                   int(round(_재료비)),                False),
                        ("",   "나. 노무비",              "",                                                   "=G11+G12",                        False),
                        ("",   "  직접노무비",            "",                                                   int(round(_직접노무비)),            False),
                        ("",   "  간접노무비",            f"직노의 {_r_간접노무:.1f}%{'(소액공사 제외)' if _소액여부 else ''}",
                                                                                                                _f_간접,                           False),
                        ("",   "다. 경비",                "",                                                   "=SUM(G14:G22)",                   False),
                        ("",   "  기계경비",              "",                                                   int(round(float(_기계경비))),       False),
                        ("",   "  기타경비",              f"(가+나)의 {_r_기타경비:.1f}%{'(소액공사 제외)' if _소액여부 else ''}",
                                                                                                                _f_기타,                           False),
                        ("",   "  산재보험료",            f"노의 {_r_산재:.2f}%",                              f"=G10*{_r_산재}/100",             False),
                        ("",   "  고용보험료",            f"노의 {_r_고용:.2f}%",                              f"=G10*{_r_고용}/100",             False),
                        ("",   "  건강보험료",            f"직노의 {_r_건강보험:.3f}%",                        f"=G11*{_r_건강보험}/100",         False),
                        ("",   "  국민연금보험료",        f"직노의 {_r_연금:.2f}%",                            f"=G11*{_r_연금}/100",             False),
                        ("",   "  장기요양보험료",        f"건강보험료의 {_r_장기요양:.2f}%",                  f"=G18*{_r_장기요양}/100",         False),
                        ("",   "  산업안전보건관리비",    f"{_안전분류} 기준",                                  int(round(_안전관리비)),            False),
                        ("",   "  환경보전비",            f"{_환경분류} / 직접공사비 기준",                    int(round(_환경보전비)),            False),
                        ("2",  "일반 관리비",             f"순공사비의 {_r_일반관리:.0f}%이내",                f"=G8*{_r_일반관리}/100",          False),
                        ("",   "소   계",                 "",                                                   "=G8+G23",                         True),
                        ("3",  "이   윤",                 f"(노+경+일관)의 {_r_이윤:.0f}%",                   f"=(G10+G13+G23)*{_r_이윤}/100",   False),
                        ("4",  "폐기물처리비",            "",                                                   int(round(_폐기물)),                False),
                        ("",   "계",                      "",                                                   "=G24+G25+G26",                    True),
                        ("5",  "부가가치세",              "계의   10%",                                        "=G27*0.1",                        False),
                        ("",   "합   계",                 "",                                                   "=G27+G28",                        True),
                    ]

                    _r = 8
                    for _no, _nm, _spec, _amt, _is_bold in _sheet_rows:
                        _ws.row_dimensions[_r].height = 16
                        _ws_set(_r, 1, _no,   align='center', size=10, border=_bdr)
                        _ws_set(_r, 2, _nm,   bold=_is_bold,  size=10, border=_bdr)
                        _ws_set(_r, 3, _spec,                 size=9,  border=_bdr)
                        _ws_set(_r, 4, "",    align='center', size=10, border=_bdr)
                        _ws_set(_r, 5, "",    align='center', size=10, border=_bdr)
                        _ws_set(_r, 6, "",    align='right',  size=10, border=_bdr)
                        _has_val = _amt is not None and (isinstance(_amt, str) or _amt > 0)
                        if _has_val:
                            _c7 = _ws_set(_r, 7, _amt if isinstance(_amt, str) else int(round(_amt)),
                                          align='right', size=10, border=_bdr)
                            _c7.number_format = '#,##0'
                            if _is_bold:
                                _c7.font = Font(name='맑은 고딕', bold=True, size=10)
                        else:
                            _ws_set(_r, 7, "", align='right', size=10, border=_bdr)
                        _ws_set(_r, 8, "", size=10, border=_bdr)
                        if _is_bold:
                            for _ci in range(1, 9):
                                _ws.cell(_r, _ci).fill = PatternFill('solid', fgColor='EEF2FF')
                        _r += 1

                    # 참고 문구
                    _r += 1
                    _ws_set(_r, 1, f'"물가자료, 물가정보 및 견적서 참조 / 조달청 간접공사비 적용기준 2026.04.13"',
                            size=9, merge_to=(_r, 8))

                    _buf = BytesIO()
                    _wb.save(_buf)
                    _buf.seek(0)
                    _fn = f"설계서용지_{(_공사명 or '원가계산')}_{datetime.today().strftime('%Y%m%d')}.xlsx"
                    st.download_button(
                        label="⬇️ 설계서용지 Excel 다운로드",
                        data=_buf.getvalue(),
                        file_name=_fn,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                except ImportError:
                    st.warning("Excel 출력에는 openpyxl이 필요합니다: `pip install openpyxl`")

        # ── 계약 방식 판단기 ────────────────────────────────────
        with _ut2:
            st.caption("📋 KIST 계약업무요령(2024.11.13 개정) 기준 · 반드시 계약팀에 사전 확인하세요.")
            if st.session_state.get("ct_price", 0) > 0:
                st.caption("📥 설계서 검토 탭에서 전달된 값이 자동 반영될 수 있습니다.")

            # ① 계약 대상
            _ct = st.radio("계약 대상", ["공사", "용역", "물품"], horizontal=True, key="ct_type")

            # ② 공사 구분
            _공사구분 = None
            if _ct == "공사":
                _공사구분 = st.radio(
                    "공사 구분 (건설산업기본법 기준)",
                    ["일반건설공사 (종합공사업 — 건축·토목 등)",
                     "전문공사 (전기·통신·소방·설비 등)",
                     "기타 공사관련 법령"],
                    horizontal=False, key="ct_consttype",
                    help="시설 인테리어·리모델링 등은 대부분 전문공사에 해당합니다"
                )

            # ③ 추정 금액
            _cp = st.number_input(
                "추정 금액 (부가세 제외, 원)",
                min_value=0, value=0, step=1_000_000, format="%d", key="ct_price"
            )

            if st.button("🔍 판단하기", type="primary", use_container_width=True, key="ct_btn") and _cp > 0:

                # ─── 공사 판단 ──────────────────────────────────
                # 원칙: 경쟁입찰. 수의계약은 제26조①5호가목 한도 이하일 때만 예외 허용.
                if _ct == "공사":
                    if "일반건설" in _공사구분:
                        _sw_limit = 400_000_000
                        _sw_label = "4억원 (일반건설공사)"
                        _type_note = "건설산업기본법 일반건설공사(전문공사 제외) 기준"
                    elif "전문공사" in _공사구분:
                        _sw_limit = 200_000_000
                        _sw_label = "2억원 (전문공사)"
                        _type_note = "건설산업기본법 전문공사 기준 — 전기·통신·소방·설비 등"
                    else:
                        _sw_limit = 160_000_000
                        _sw_label = "1억6천만원 (기타)"
                        _type_note = "기타 공사관련 법령 적용 공사 기준"

                    if _cp <= 20_000_000:
                        _cm = "수의계약 가능 (소액 · 1인 견적)"; _cc = "#2f9e44"
                        _cd = f"2,000만원 이하 · 소액 수의계약 예외 적용 (제26조①5호나목2)"
                        _docs = [
                            "견적서 1개 이상",
                            "설계서 (내역서, 도면, 시방서) — 해당 시",
                            "수의계약 체결제한여부확인서 (제26조의2)",
                        ]
                        _notes = [
                            "근거: 계약업무요령 제26조①5호나목2 · 제29조①2호",
                            "⚠️ 수의계약은 예외 — 특별한 사유 없이 관행적으로 사용 금지",
                            "분할 발주 금지 (동일 목적 사업 합산 적용)",
                            "계약상대자 이해충돌방지법 확인 필수",
                        ]
                        _sw_box = None
                    elif _cp <= _sw_limit:
                        _cm = "수의계약 가능 (2인 이상 견적 + 나라장터 공고)"; _cc = "#f08c00"
                        _cd = f"2,000만원 초과 ~ {_sw_label} 이하 · 수의계약 예외 적용 가능"
                        _docs = [
                            "견적서 2개 이상 (동일 규격·수량·조건)",
                            "설계도서 (내역서, 도면, 시방서)",
                            "원가계산서 (설계금액 기준)",
                            "수의계약 체결제한여부확인서 (제26조의2)",
                            "나라장터 안내공고 (전자조달시스템)",
                        ]
                        _notes = [
                            f"근거: 계약업무요령 제26조①5호가목1 · 제29조②",
                            f"{_type_note}",
                            "⚠️ 금액 한도 내라도 경쟁입찰이 원칙 — 수의계약은 예외적 선택",
                            "나라장터 안내공고 후 견적서 접수 (제29조②)",
                            "동일 조건 2인 이상 견적 비교 후 계약",
                            "분할 발주 및 특정 업체 지정 금지",
                            "계약팀 사전 협의 권장",
                        ]
                        _sw_box = None
                    else:
                        _cm = "경쟁입찰"; _cc = "#c92a2a"
                        _cd = f"{_sw_label} 초과 · 경쟁입찰 (수의계약 불가)"
                        _docs = [
                            "설계도서 (내역서, 도면, 시방서)",
                            "원가계산서",
                            "입찰공고문 (나라장터 공고)",
                            "현장설명서 (해당 시)",
                            "입찰참가자격 서류",
                        ]
                        _notes = [
                            "근거: 계약업무요령 제14조·제23조",
                            f"{_type_note}",
                            "나라장터(KONEPS) 공고 원칙",
                            "추정가격 10억 미만: 공고 7일 전",
                            "추정가격 10억~50억 미만: 공고 15일 전",
                            "추정가격 50억 이상: 공고 40일 전 (제32조)",
                            "적격심사 또는 최저가 낙찰제 적용",
                            "계약팀 사전 협의 필수",
                        ]
                        _sw_box = None

                # ─── 용역·물품 판단 ─────────────────────────────
                # 원칙: 경쟁입찰. 2천만원 이하만 소액 수의 예외.
                # 2천만원 초과는 일괄 경쟁입찰 원칙 — 수의 예외 사유는 별도 안내.
                else:
                    _is_용역 = (_ct == "용역")
                    if _cp <= 20_000_000:
                        _cm = "수의계약 가능 (소액 · 1인 견적)"; _cc = "#2f9e44"
                        _cd = f"2,000만원 이하 · 소액 수의계약 예외 적용 (제26조①5호나목2)"
                        if _is_용역:
                            _docs = [
                                "견적서 1개 이상",
                                "과업지시서 (또는 용역범위 기술서)",
                                "수의계약 체결제한여부확인서 (제26조의2)",
                            ]
                        else:
                            _docs = [
                                "견적서 1개 이상",
                                "규격서 / 물품 사양서",
                                "수의계약 체결제한여부확인서 (제26조의2)",
                            ]
                        _notes = [
                            "근거: 계약업무요령 제26조①5호나목2 · 제29조①2호",
                            "⚠️ 수의계약은 예외 — 특별한 사유 없이 관행적으로 사용 금지",
                            "분할 발주 금지 (동일 목적 사업 합산 적용)",
                            "계약상대자 이해충돌방지법 확인 필수",
                        ]
                        if not _is_용역:
                            _notes.append("'동등 이상' 허용 문구 반드시 포함 / 특정 상표·모델 고정 금지")
                        _sw_box = None

                    else:
                        # 2천만원 초과 → 경쟁입찰 원칙
                        _cm = "경쟁입찰"; _cc = "#c92a2a"
                        _cd = "2,000만원 초과 · 경쟁입찰 원칙 (수의계약은 별도 예외 사유 필요)"
                        if _is_용역:
                            _docs = [
                                "과업지시서 / 제안요청서(RFP)",
                                "원가계산서",
                                "입찰공고문 (나라장터)",
                                "제안서 평가 기준 (협상방식 시)",
                            ]
                            _notes = [
                                "근거: 계약업무요령 제14조·제23조·제43조",
                                "나라장터(KONEPS) 공고 원칙",
                                "1억원 초과: 협상에 의한 계약(기술·가격 분리) 가능 (제21조·제43조)",
                                "계약팀 사전 협의 필수",
                            ]
                        else:
                            _docs = [
                                "규격서 / 물품 사양서",
                                "입찰공고문 (나라장터) 또는 MAS 구매요구서",
                                "원가계산서 (해당 시)",
                            ]
                            _notes = [
                                "근거: 계약업무요령 제14조·제23조",
                                "나라장터(KONEPS) 공고 원칙",
                                "조달청 MAS(다수공급자계약) 우선 활용 검토",
                                "'동등 이상' 허용 문구 반드시 포함",
                                "계약팀 사전 협의 필수",
                            ]
                        # 수의계약 예외 조건 안내 박스
                        _sw_box = (
                            "**이 금액대에서 수의계약이 가능한 예외 사유 (계약업무요령 제26조①5호)**\n\n"
                            "아래 조건 중 하나에 해당하는 경우, 계약팀과 협의 후 수의계약 가능:\n\n"
                            "- **나목③** 소기업·소상공인과의 계약으로 추정가격이 **1억원 이하** (소기업·소상공인 확인서 필요)\n"
                            "- **나목④** 학술연구·원가계산·건설기술 등 **특수 지식·기술·자격**을 요구하는 용역 (1억원 이하)\n"
                            "- **나목⑤** 여성기업·장애인기업·사회적기업 등과의 계약 (1억원 이하)\n"
                            "- **제26조①2호** 특정인 기술·단독 공급·특허 등 경쟁 성립 불가 사유\n\n"
                            "⚠️ 위 예외를 적용하더라도 반드시 **수의계약 사유서** 작성 및 **계약팀 사전 승인** 필요"
                        )

                # ─── 결과 출력 ───────────────────────────────────
                st.markdown(
                    f"<div style='padding:22px 20px;background:{_cc};border-radius:12px;"
                    f"text-align:center;margin:16px 0;'>"
                    f"<div style='font-size:28px;font-weight:800;color:white;margin-bottom:6px;'>{_cm}</div>"
                    f"<div style='font-size:13px;color:rgba(255,255,255,0.88);margin-bottom:6px;'>{_cd}</div>"
                    f"<div style='font-size:17px;font-weight:700;color:white;'>"
                    f"추정금액: {_cp:,}원 (부가세 제외)</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

                _da, _db = st.columns(2)
                with _da:
                    st.subheader("📄 필요 서류")
                    for _d in _docs:
                        st.markdown(f"✅ {_d}")
                with _db:
                    st.subheader("⚠️ 근거 및 주의사항")
                    for _n in _notes:
                        st.markdown(f"🔸 {_n}")

                # 수의계약 예외 조건 안내 (2천만원 초과 용역·물품)
                if _sw_box:
                    with st.expander("💡 수의계약 예외 사유 해당 여부 확인 (클릭)", expanded=False):
                        st.markdown(_sw_box)

                # 견적서 제출 요건 안내 (제29조)
                with st.expander("📌 견적서 제출 요건 안내 (계약업무요령 제29조)", expanded=False):
                    st.markdown("""
| 추정가격 | 견적서 수 | 전자조달시스템 |
|---------|---------|------------|
| **2천만원 이하** | 1인 이상 | 불필요 |
| **2천만원 초과 수의계약** | **2인 이상** | **나라장터 안내공고 필수** |

- 견적서는 동일한 규격·수량·조건으로 요청하여야 합니다.
- 나라장터 안내공고 결과 1인만 제출 시, 재공고 후에도 1인이면 1인 견적으로 진행 가능.
- 견적가격이 예정가격 범위를 초과하면 재견적 요청.
                    """)

                st.markdown("""
<style>
.ct-tip{position:relative;display:inline-block;color:#1a73e8;font-weight:700;
  border-bottom:2px dashed #1a73e8;cursor:help;}
.ct-box{visibility:hidden;opacity:0;width:360px;background:#1e293b;color:#e2e8f0;
  font-size:12px;line-height:1.75;padding:13px 15px;border-radius:9px;
  position:absolute;bottom:140%;left:50%;transform:translateX(-50%);
  z-index:9999;transition:opacity .2s;box-shadow:0 6px 22px rgba(0,0,0,.4);
  white-space:normal;}
.ct-box::after{content:"";position:absolute;top:100%;left:50%;
  margin-left:-6px;border:6px solid transparent;border-top-color:#1e293b;}
.ct-tip:hover .ct-box{visibility:visible;opacity:1;}
</style>
<div style="background:#e8f4fd;border-left:4px solid #1a73e8;padding:10px 15px;
  border-radius:6px;font-size:13px;color:#1a3a5c;line-height:1.8;">
※ 계약의 원칙은 <b>경쟁입찰</b>입니다. 수의계약은
<span class="ct-tip">제26조 예외 사유
  <span class="ct-box">
    <b style="color:#93c5fd;">📋 수의계약 주요 예외 사유 (제26조①)</b><br>
    <span style="color:#fbbf24;">1호</span> 긴급·비상재해·보안 등 경쟁 불가<br>
    <span style="color:#fbbf24;">2호</span> 특허·단독공급·특정기술 등 경쟁 성립 불가<br>
    <span style="color:#fbbf24;">5호 나목②</span> 소액 2천만원 이하<br>
    <span style="color:#fbbf24;">5호 나목③</span> 소기업·소상공인 — 1억원 이하<br>
    <span style="color:#fbbf24;">5호 나목④</span> 특수기술·학술연구 용역 — 1억원 이하<br>
    <span style="color:#fbbf24;">5호 나목⑤</span> 여성기업·장애인기업·사회적기업 — 1억원 이하<br>
    <span style="color:#fbbf24;">5호 가목①</span> 공사 한도 이하 (일반건설 4억 / 전문공사 2억 / 기타 1.6억)<br>
    <span style="color:#86efac;font-size:11px;">※ 위 사유 해당 시에도 계약팀 사전 협의 필수</span>
  </span>
</span>
에 해당할 때만 가능합니다. 반드시 계약팀에 사전 확인하세요.
</div>
""", unsafe_allow_html=True)

            elif _cp == 0:
                st.markdown(
                    "<div style='text-align:center;padding:40px;color:#aaa;font-size:14px;'>"
                    "추정 금액을 입력하고 [판단하기] 버튼을 누르세요.</div>",
                    unsafe_allow_html=True,
                )

            # 하단 금액 기준표 요약
            with st.expander("📊 KIST 계약방식 금액 기준 요약표 (계약업무요령 제26조)", expanded=False):
                st.markdown("""
> **⚠️ 계약의 원칙은 경쟁입찰입니다. 수의계약은 아래 요건에 해당할 때만 예외적으로 가능합니다.**

**공사 수의계약 허용 한도 (제26조①5호가목1)**

| 공사 구분 | 수의계약 허용 한도 | 한도 초과 시 |
|---------|--------------|-----------|
| 일반건설공사 (종합) | 추정가격 **4억원 이하** | 경쟁입찰 |
| 전문공사 (전기·통신·소방·설비 등) | 추정가격 **2억원 이하** | 경쟁입찰 |
| 기타 공사관련 법령 공사 | 추정가격 **1억6천만원 이하** | 경쟁입찰 |

※ 한도 이하라도 경쟁입찰이 원칙이며, 수의계약은 예외적 선택입니다.

**물품·용역 수의계약 허용 범위 (제26조①5호나목)**

| 조건 | 수의계약 허용 한도 | 근거 |
|-----|--------------|-----|
| 일반 (소액) | 추정가격 **2천만원 이하** | 나목② |
| 소기업·소상공인 | 추정가격 **1억원 이하** | 나목③ |
| 특수기술·학술연구 용역 | 추정가격 **1억원 이하** | 나목④ |
| 여성기업·장애인기업·사회적기업 | 추정가격 **1억원 이하** | 나목⑤ |

※ **2천만원 초과는 경쟁입찰이 원칙**이며, 소기업·특수기술 등 예외 사유는 별도 사유서 및 계약팀 승인 필요.

> 위 금액은 **추정가격(부가세 제외)** 기준이며, **분할 발주는 합산 적용**됩니다.
                """)

        # ── 소액공사 서류 생성기 ───────────────────────────────
        with _ut3:
            _template_path = find_small_work_template()
            if _template_path:
                st.caption(f"📄 양식 파일 연결됨: {_template_path}")
            else:
                st.warning("양식 파일을 찾지 못했습니다. Desktop/소액공사 자료/승락서,지급각서,하자각서.xlsx 위치를 확인하세요.")

            st.caption("승락서 · 지급각서 · 하자보수책임각서 · 청렴계약 이행각서를 생성합니다.")

            _swc1, _swc2 = st.columns(2)
            with _swc1:
                _sw_project_name = st.text_input("공사명", placeholder="예: 본관 A361호 공조덕트 보수공사", key="sw_project_name")
                st.markdown(
                    """
<div style='display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;'>
  <div style='font-size:0.95rem;font-weight:600;color:rgb(38, 39, 48);'>업체명</div>
  <a href='http://p.kist.re.kr:8081/nxui/kistis/indexQ.jsp?target=mis.pur::pur_1310.xfdl&menuParam=sysCd=CUS'
     target='_blank'
     style='font-size:0.78rem;color:#1a73e8;text-decoration:none;white-space:nowrap;'>업체등록 또는 정보조회</a>
</div>
                    """,
                    unsafe_allow_html=True,
                )
                _sw_company = st.text_input("업체명 입력", placeholder="예: 유진건설", key="sw_company", label_visibility="collapsed")
                _sw_ceo = st.text_input("대표자", placeholder="예: 도현수", key="sw_ceo")
                _sw_address = st.text_area("주소", placeholder="예: 서울시 성북구 화랑로 1026-0708번지", height=100, key="sw_address")
                _sw_seal_file = st.file_uploader("직인 이미지 업로드", type=["png", "jpg", "jpeg", "webp"], key="sw_seal_file")
            with _swc2:
                _sw_contract_date = st.date_input("계약일자", value=datetime.today(), key="sw_contract_date")
                _sw_start_date = st.date_input("공사 시작일", value=datetime.today(), key="sw_start_date")
                _sw_end_date = st.date_input("공사 종료일", value=datetime.today(), key="sw_end_date")
                _sw_base_amount = st.number_input("통제금액 / 기초금액 (원)", min_value=0, value=0, step=100_000, format="%d", key="sw_base_amount")
                _sw_contract_amount = st.number_input("계약금액 (원, 부가세 포함)", min_value=0, value=0, step=100_000, format="%d", key="sw_contract_amount")
                _sw_contract_method = st.text_input("계약방법", value="수의계약", key="sw_contract_method")
                _sw_defect_label = st.selectbox(
                    "하자보증금율",
                    ["건설업종 3%", "조경 5%", "전기, 통신, 소방 등 건설업종 외의 공사 2%"],
                    index=0,
                    key="sw_defect_label",
                )
                _sw_defect_period = st.text_input("하자보수책임기간", value="2년", key="sw_defect_period")

            with st.expander("생성 내용 미리 보기", expanded=True):
                st.markdown(
                    f"""
- 공사명: {_sw_project_name or '-'}
- 계약일자: {_sw_contract_date.strftime('%Y-%m-%d')}
- 공사기간: {_sw_start_date.strftime('%Y-%m-%d')} ~ {_sw_end_date.strftime('%Y-%m-%d')}
- 계약업체명: {_sw_company or '-'}
- 대표자: {_sw_ceo or '-'}
- 주소: {_sw_address or '-'}
- 통제금액/기초금액: {_sw_base_amount:,}원
- 계약금액: {_sw_contract_amount:,}원
- 계약방법: {_sw_contract_method or '-'}
- 하자보증금율: {_sw_defect_label}
- 하자보수책임기간: {_sw_defect_period or '-'}
- 직인 이미지: {'업로드됨' if _sw_seal_file else '없음'}
                    """
                )

            _missing = []
            if not _sw_project_name.strip():
                _missing.append("공사명")
            if not _sw_company.strip():
                _missing.append("업체명")
            if not _sw_ceo.strip():
                _missing.append("대표자")
            if not _sw_address.strip():
                _missing.append("주소")
            if _sw_contract_amount <= 0:
                _missing.append("계약금액")
            if _sw_end_date < _sw_start_date:
                st.error("공사 종료일은 시작일보다 빠를 수 없습니다.")
                _missing.append("공사기간")

            if _missing:
                st.caption("입력 필요: " + ", ".join(_missing))

            _btn1, _btn2, _btn3 = st.columns(3)
            with _btn1:
                _sw_generate_clicked = st.button("📄 서류 생성", type="primary", use_container_width=True, key="sw_generate_btn")
            with _btn2:
                st.button("PDF 자동 포함", use_container_width=True, disabled=True, key="sw_pdf_btn")
            with _btn3:
                st.button("계약요청 문구 복사", use_container_width=True, disabled=True, key="sw_copy_btn")

            if _sw_generate_clicked:
                if _missing:
                    st.error("필수 입력값을 먼저 확인하세요: " + ", ".join(_missing))
                else:
                    try:
                        _small_work_data = {
                            "contract_date": _sw_contract_date,
                            "start_date": _sw_start_date,
                            "end_date": _sw_end_date,
                            "project_name": _sw_project_name.strip(),
                            "base_amount": _sw_base_amount,
                            "contract_amount": _sw_contract_amount,
                            "contract_method": _sw_contract_method.strip() or "수의계약",
                            "company_name": _sw_company.strip(),
                            "ceo_name": _sw_ceo.strip(),
                            "address": _sw_address.strip(),
                            "defect_label": _sw_defect_label,
                            "defect_rate": {
                                "건설업종 3%": 0.03,
                                "조경 5%": 0.05,
                                "전기, 통신, 소방 등 건설업종 외의 공사 2%": 0.02,
                            }[_sw_defect_label],
                            "defect_period": _sw_defect_period.strip() or "2년",
                            "seal_image_bytes": _sw_seal_file.getvalue() if _sw_seal_file else None,
                        }
                        _xlsx_bytes, _xlsx_name = build_small_work_doc_xlsx(_small_work_data)
                        _pdf_bytes, _pdf_name = build_integrity_pledge_pdf(_small_work_data)

                        _zip_buf = BytesIO()
                        with zipfile.ZipFile(_zip_buf, "w", zipfile.ZIP_DEFLATED) as _zip:
                            _zip.writestr(_xlsx_name, _xlsx_bytes)
                            _zip.writestr(_pdf_name, _pdf_bytes)
                        _zip_buf.seek(0)
                        _zip_name = f"소액공사서류_{re.sub(r'[\\/:*?\"<>|]+', '_', _sw_project_name.strip())[:80]}_{_sw_contract_date.strftime('%Y%m%d')}.zip"

                        st.success("양식 파일 생성 완료")
                        st.download_button(
                            label="⬇️ 전체 서류 ZIP 다운로드",
                            data=_zip_buf.getvalue(),
                            file_name=_zip_name,
                            mime="application/zip",
                            use_container_width=True,
                            key="sw_download_zip_btn",
                        )
                        st.download_button(
                            label="⬇️ Excel만 다운로드",
                            data=_xlsx_bytes,
                            file_name=_xlsx_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="sw_download_xlsx_btn",
                        )
                        st.download_button(
                            label="⬇️ 청렴계약 이행각서 PDF만 다운로드",
                            data=_pdf_bytes,
                            file_name=_pdf_name,
                            mime="application/pdf",
                            use_container_width=True,
                            key="sw_download_pdf_btn",
                        )
                    except Exception as _e:
                        st.error(f"서류 생성 오류: {_e}")

            st.caption("참고: 원본 양식의 수식을 유지하며, 지급각서의 외부참조 수식 2개와 하자각서 날짜는 생성 시 자동 보정합니다.")


    # ── Tab 4: 부서정보 편집 ──────────────────────────────────
    with tab4:
        st.subheader("부서정보.md 편집")
        st.caption("이 내용이 AI 보고서 생성 시 자동으로 반영됩니다.")

        ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")
        if not ADMIN_PASSWORD:
            try:
                ADMIN_PASSWORD = st.secrets["ADMIN_PASSWORD"]
            except Exception:
                ADMIN_PASSWORD = "kist2026"

        if "admin_ok" not in st.session_state:
            st.session_state.admin_ok = False

        if not st.session_state.admin_ok:
            pw = st.text_input("관리자 비밀번호", type="password", placeholder="비밀번호 입력 후 Enter")
            if pw:
                if pw == ADMIN_PASSWORD:
                    st.session_state.admin_ok = True
                    st.rerun()
                else:
                    st.error("비밀번호가 틀렸습니다.")
        else:
            if os.path.exists(CONTEXT_FILE):
                with open(CONTEXT_FILE, "r", encoding="utf-8") as f:
                    current_content = f.read()
            else:
                current_content = ""

            edited_content = st.text_area(
                label="부서정보",
                value=current_content,
                height=550,
                label_visibility="collapsed",
            )

            col_a, col_b, col_c = st.columns([1, 1, 4])
            with col_a:
                if st.button("💾 저장", type="primary", use_container_width=True):
                    with open(CONTEXT_FILE, "w", encoding="utf-8") as f:
                        f.write(edited_content)
                    st.success("저장 완료!")
            with col_b:
                if st.button("🔒 잠금", use_container_width=True):
                    st.session_state.admin_ok = False
                    st.rerun()


# ══════════════════════════════════════════════════════════════
# 우측: 원규·지침 검색 패널 (항상 표시)
# ══════════════════════════════════════════════════════════════
with _right_col:
    st.markdown("### 📚 원규·지침 검색")

    # DB 상태 배지
    _db_ok = False
    try:
        from kist_rag import db_stats
        _s = db_stats()
        _db_ok = _s["ready"]
        if _db_ok:
            st.caption(f"🟢 {_s['total_chunks']:,}개 조항 인덱싱됨")
        else:
            st.caption("🔴 DB 미구축 — build_index.py 실행 필요")
    except Exception:
        st.caption("🔴 chromadb 미설치 또는 DB 없음")

    st.divider()

    # 검색 폼
    with st.form("rag_search_form", clear_on_submit=False):
        _query = st.text_input(
            "검색어",
            value=st.session_state.rag_query,
            placeholder="예: 수의계약 한도금액",
            label_visibility="collapsed",
        )
        _cat = st.radio(
            "범위",
            ["전체", "원규", "지침"],
            horizontal=True,
            label_visibility="collapsed",
        )
        _search_btn = st.form_submit_button("🔍 검색", use_container_width=True, disabled=not _db_ok)

    if _search_btn and _query.strip():
        try:
            from kist_rag import search_hybrid
            _cat_map = {"전체": None, "원규": "원규정보", "지침": "지침정보"}
            _items = search_hybrid(_query.strip(), n_results=6, category=_cat_map[_cat])
            st.session_state.rag_results = _items
            st.session_state.rag_query = _query.strip()
        except Exception as e:
            st.error(f"검색 오류: {e}")

    # 검색 결과
    if st.session_state.rag_results:
        from kist_rag import highlight as rag_highlight
        _kws = [w for w in st.session_state.rag_query.split() if len(w) >= 2]

        st.markdown(
            f"<div style='font-size:0.82rem;color:#555;margin-bottom:4px;'>"
            f"<b>{st.session_state.rag_query}</b> 검색 결과 {len(st.session_state.rag_results)}건</div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            "<div style='font-size:0.78rem;color:#e67e22;background:#fff8f0;"
            "border-left:3px solid #e67e22;padding:5px 9px;border-radius:4px;margin-bottom:8px;'>"
            "⚠️ 단서 참고용 — 정확한 내용은 반드시 스마트규정시스템 원문으로 확인하세요."
            "</div>",
            unsafe_allow_html=True,
        )
        for i, item in enumerate(st.session_state.rag_results):
            _src = item["source"].replace("[원규정보] ", "").replace("[지침정보] ", "")
            _badge = "📘" if "원규" in item["source"] else "📗"
            _art = item.get("article", "")

            # 순수 규정명 추출: 괄호·연도·개정전문 등 제거
            import urllib.parse
            _doc_name = re.sub(r'\s*\(.*?\)', '', _src)          # (2024년도 11월 개정) 제거
            _doc_name = re.sub(r'\s*(개정전문|제정전문|전문개정|개정|제정)\s*$', '', _doc_name)  # 말미 접미사
            _doc_name = re.sub(r'^\(?\d{6}\)?\s*', '', _doc_name)  # (210923) 같은 날짜코드
            _doc_name = _doc_name.replace('.hwp', '').strip()
            _law_url = (
                "http://law.kist.re.kr:8090/lmxsrv/law/lawList.do"
                f"?searchType=title&searchText={urllib.parse.quote(_doc_name)}"
            )

            with st.expander(f"{_badge} {_src[:30]}", expanded=(i == 0)):
                if _art:
                    st.markdown(
                        f"<span style='background:#e8f4fd;color:#1a73e8;font-size:0.8rem;"
                        f"padding:2px 7px;border-radius:10px;font-weight:600;'>{_art}</span>",
                        unsafe_allow_html=True,
                    )
                    st.write("")

                _body = rag_highlight(item["text"], _kws)
                st.markdown(
                    f"<div style='font-size:0.86rem;line-height:1.75;color:#222;'>{_body}</div>",
                    unsafe_allow_html=True,
                )

                st.markdown("---")
                _main_url = "http://law.kist.re.kr:8090/lmxsrv/main/main.srv"
                st.markdown(
                    f"""<a href="{_main_url}" target="_blank"
                        style="display:block;text-align:center;padding:7px;
                               background:#1a73e8;color:white;border-radius:6px;
                               text-decoration:none;font-size:13px;">
                        🔗 스마트규정시스템 열기
                    </a>""",
                    unsafe_allow_html=True,
                )
                # 복사용 문서명: 원본 source 그대로 사용 (파일명 기반)
                # 스마트규정시스템은 띄어쓰기가 다를 수 있으므로 안내 표시
                _copy_label = _src.strip()   # 원본 source (카테고리 접두사 제거된 것)
                st.markdown(
                    f"<div style='margin-top:6px;'>"
                    f"<div style='font-size:11px;color:#888;margin-bottom:3px;'>📋 스마트규정시스템 검색어 (복사 후 붙여넣기)</div>"
                    f"<div style='display:flex;align-items:center;gap:6px;'>"
                    f"<div style='flex:1;padding:6px 10px;"
                    f"background:#f8f9fa;border:1px solid #dee2e6;border-radius:6px;"
                    f"font-size:13px;color:#333;font-weight:600;user-select:all;cursor:text;'>"
                    f"{_doc_name}</div></div>"
                    f"<div style='font-size:11px;color:#e67e22;margin-top:4px;'>"
                    f"⚠️ HWP 파일명 기준 — 시스템 내 실제 문서명과 띄어쓰기가 다를 수 있습니다.<br>"
                    f"&nbsp;&nbsp;&nbsp;&nbsp;검색 안 되면 앞 2~3 글자만 입력해 보세요. (예: <b>연봉제</b>)"
                    f"</div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )

    elif st.session_state.rag_query:
        st.info("검색 결과가 없습니다.")
    else:
        if _db_ok:
            st.markdown(
                "<div style='color:#888;font-size:0.85rem;margin-top:8px;'>"
                "업무 중 궁금한 규정을<br>바로 검색하세요.<br><br>"
                "예시 검색어:<br>"
                "· 수의계약 한도<br>"
                "· 시설공사 원가계산<br>"
                "· 안전관리비 계상<br>"
                "· 견적서 징구 기준"
                "</div>",
                unsafe_allow_html=True,
            )

st.divider()
st.caption("제작 : 김진광 전문원")
